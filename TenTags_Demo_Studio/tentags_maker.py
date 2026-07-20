import ast
import os
import queue
import subprocess
import sys
import tempfile
import threading
import tkinter as tk
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from tkinter import colorchooser, filedialog, messagebox, ttk
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen
import tkinter.font as tkfont

import tentags


APP_DIR = Path(__file__).resolve().parent

DEFAULT_PREAMBLE = '4,7,1,"#cbd5e1","solid-1",1,48'

DEFAULT_STYLE = """style(
<cm><bg=#1e293b><color=#f8fafc><b><fs=20px>, , , , , , </fs></b></color></bg></cm>;
<center><bg=#f1f5f9><color=#0f172a><b><fs=16px></fs></b></color></bg></center>, <center><bg=#e2e8f0><color=#0f172a><b><fs=16px></fs></b></color></bg></center>, <center><bg=#cbd5e1><color=#0f172a><b><fs=16px></fs></b></color></bg></center>, <center><bg=#94a3b8><color=#f8fafc><b><fs=16px></fs></b></color></bg></center>, <center><bg=#64748b><color=#f8fafc><b><fs=16px></fs></b></color></bg></center>, <center><bg=#475569><color=#f8fafc><b><fs=16px></fs></b></color></bg></center>, <center><bg=#334155><color=#f8fafc><b><fs=16px></fs></b></color></bg></center>;
<cm><bg=#f8fafc><color=#64748b><i><fs=11px></fs></i></color></bg></cm>, , , <center><bg=#ffffff></bg></center>, <cm><bg=#f8fafc><color=#64748b><i><fs=11px></fs></i></color></bg></cm>, , ;
<cm><bg=#eff6ff><color=#1d4ed8><b><u><fs=14px>, , , , , , </fs></u></b></color></bg></cm>
)"""

DEFAULT_DATA = """data(
<cm>TenTags Studio, , , , , , </cm>;
T, e, n, T, a, g, s;
<cm>Modern Python, , </cm>, <img src=https://tentags.org/assets/img/tentags_logo.png w=30 h=auto m=12>, <cm>Reporting Engine, , </cm>;
<cm><url=https://pycells.com>Powering PyCells Tables <img src=https://pycells.com/assets/img/PyCells_mds.png w=20 h=auto m=8></url>, , , , , , </cm>
)"""


def choose_font_family(root, candidates):

    available = {
        family.casefold(): family
        for family in tkfont.families(root)
    }
    for candidate in candidates:
        family = available.get(candidate.casefold())
        if family:
            return family
    return candidates[-1]


# ============================================================
# Cell model and widget
# ============================================================


@dataclass
class CellState:
    text: str = ""
    bold: bool = False
    italic: bool = False
    underline: bool = False
    strike: bool = False
    font_size: str = ""
    align: str = "Left"
    url: str = ""
    mark: str = ""
    image: str = ""
    image_w: str = "120"
    image_h: str = "auto"
    image_m: str = "15"
    bg: str = "#ffffff"
    fg: str = "#000000"
    hide_left: bool = False
    hide_right: bool = False
    hide_top: bool = False
    hide_bottom: bool = False


class ImagePreviewCache:

    def __init__(self, owner, ready_callback):

        self.owner = owner
        self.ready_callback = ready_callback
        self.cache = {}
        self.loading = set()
        self.results = queue.Queue()
        self.owner.after(100, self._poll)

    def get(self, source):

        source = str(source or "").strip()
        if not source:
            return "error", None
        cached = self.cache.get(source)
        if cached is not None:
            return cached
        if source not in self.loading:
            self.loading.add(source)
            threading.Thread(
                target=self._load,
                args=(source,),
                daemon=True
            ).start()
        return "loading", None

    @staticmethod
    def _local_path(source):

        decoded = unquote(source)
        if decoded.lower().startswith("file://"):
            parsed = urlparse(decoded)
            decoded = parsed.path
            if parsed.netloc:
                decoded = f"//{parsed.netloc}{decoded}"
            if len(decoded) >= 3 and decoded[0] == "/" and decoded[2] == ":":
                decoded = decoded[1:]
        path = Path(decoded)
        if not path.is_absolute():
            path = APP_DIR / path
        return path.resolve()

    def _load(self, source):

        try:
            from PIL import Image, ImageOps

            if source.lower().startswith(("http://", "https://")):
                request = Request(
                    source,
                    headers={"User-Agent": "TenTags-Studio/1.0"}
                )
                with urlopen(request, timeout=20) as response:
                    payload = BytesIO(response.read())
                with Image.open(payload) as opened:
                    image = ImageOps.exif_transpose(opened).convert("RGBA")
            else:
                path = self._local_path(source)
                if not path.is_file():
                    raise FileNotFoundError(path)
                with Image.open(path) as opened:
                    image = ImageOps.exif_transpose(opened).convert("RGBA")
            self.results.put((source, "ready", image.copy()))
        except Exception as exc:
            self.results.put((source, "error", str(exc)))

    def _poll(self):

        try:
            while True:
                source, status, value = self.results.get_nowait()
                self.loading.discard(source)
                self.cache[source] = (status, value)
                self.ready_callback(source)
        except queue.Empty:
            pass

        try:
            exists = self.owner.winfo_exists()
        except tk.TclError:
            exists = False
        if exists:
            self.owner.after(100, self._poll)


class Cell(tk.Frame):

    CELL_WIDTH = 110
    CELL_HEIGHT = 36

    def __init__(
        self,
        master,
        row,
        col,
        state,
        edit_callback=None,
        cell_width=None,
        cell_height=None,
        border_width=1,
        border_color="#9ca3af",
        border_style="solid-1",
        outer_edges=None,
        image_provider=None
    ):

        self.cell_width = cell_width or self.CELL_WIDTH
        self.cell_height = cell_height or self.CELL_HEIGHT
        self.preview_border_width = border_width
        self.preview_border_color = border_color
        self.preview_border_style = border_style
        self.outer_edges = set(outer_edges or ())
        self.image_provider = image_provider

        super().__init__(
            master,
            width=self.cell_width,
            height=self.cell_height,
            background=state.bg,
            borderwidth=0,
            highlightthickness=0,
            takefocus=True
        )

        self.grid_propagate(False)

        self.row = row
        self.col = col
        self.state = state
        self.selected = False
        self.edit_callback = edit_callback
        self.inline_editor = None
        self.preview_photo = None

        self.label = tk.Label(
            self,
            text="",
            bg=self.state.bg,
            fg=self.state.fg,
            anchor="w",
            padx=5,
            borderwidth=0,
            highlightthickness=0
        )

        self.label.place(x=1, y=1, relwidth=1, relheight=1, width=-2, height=-2)

        self.edges = {
            name: tk.Canvas(
                self,
                background=self.state.bg,
                borderwidth=0,
                highlightthickness=0
            )
            for name in ("top", "right", "bottom", "left")
        }

        self.bind("<Button-1>", self.on_click)
        self.label.bind("<Button-1>", self.on_click)
        self.bind("<Double-Button-1>", self.on_double_click)
        self.label.bind("<Double-Button-1>", self.on_double_click)
        self.bind("<F2>", self.start_edit)
        self.bind("<Return>", self.start_edit)
        self.bind("<KeyPress>", self.on_key_press)
        self.refresh()

    def on_click(self, event=None):

        self.focus_set()
        self.event_generate("<<CellSelected>>")

    def on_double_click(self, event=None):

        self.on_click()
        self.after_idle(self.start_edit)
        return "break"

    def on_key_press(self, event):

        if self.inline_editor is not None:
            return None
        if event.keysym in {"F2", "Return", "Escape", "Tab"}:
            return None
        if event.char and event.char.isprintable() and not (event.state & 0xC):
            self.start_edit(initial_text=event.char, replace=True)
            return "break"
        return None

    def start_edit(self, event=None, initial_text=None, replace=False):

        if self.inline_editor is not None:
            return "break"

        self.on_click()
        editor = tk.Entry(
            self,
            borderwidth=1,
            relief="solid",
            bg=self.state.bg,
            fg=self.state.fg,
            insertbackground=self.state.fg,
            font=self.label.cget("font")
        )
        self.inline_editor = editor
        editor.insert(0, initial_text if replace else self.state.text)
        editor.place(x=2, y=2, relwidth=1, relheight=1, width=-4, height=-4)
        editor.lift()
        editor.focus_set()
        if not replace:
            editor.selection_range(0, "end")

        editor.bind("<Return>", self.commit_edit)
        editor.bind("<Escape>", self.cancel_edit)
        editor.bind("<FocusOut>", self.commit_edit)
        return "break"

    def commit_edit(self, event=None):

        editor = self.inline_editor
        if editor is None:
            return "break"

        value = editor.get()
        self.inline_editor = None
        editor.destroy()
        self.state.text = value
        self.refresh()
        self.focus_set()
        if self.edit_callback:
            self.edit_callback(self)
        return "break"

    def cancel_edit(self, event=None):

        editor = self.inline_editor
        if editor is not None:
            self.inline_editor = None
            editor.destroy()
            self.focus_set()
        return "break"

    def select(self):

        self.selected = True
        self.label.configure(
            highlightthickness=2,
            highlightbackground="#2563eb",
            highlightcolor="#2563eb"
        )

    def deselect(self):

        self.selected = False
        self.label.configure(highlightthickness=0)

    def set_text(self, text):

        self.state.text = text
        self.refresh()

    def set_background(self, color):

        self.state.bg = color
        self.refresh()

    def set_foreground(self, color):

        self.state.fg = color
        self.refresh()

    def set_preview_geometry(
        self,
        width,
        height,
        border_width,
        border_color,
        border_style
    ):

        self.cell_width = max(1, int(width))
        self.cell_height = max(1, int(height))
        self.preview_border_width = max(0, int(border_width))
        self.preview_border_color = border_color
        self.preview_border_style = border_style
        self.configure(width=self.cell_width, height=self.cell_height)
        self.refresh()

    @staticmethod
    def _image_dimension(value):

        if value is None or str(value).strip().lower() == "auto":
            return None
        try:
            return max(1.0, float(value))
        except (TypeError, ValueError):
            return None

    def _build_preview_photo(self, image):

        from PIL import Image, ImageTk

        original_width, original_height = image.size
        if original_width <= 0 or original_height <= 0:
            return None

        requested_width = self._image_dimension(self.state.image_w)
        requested_height = self._image_dimension(self.state.image_h)
        if requested_width is not None and requested_height is not None:
            width, height = requested_width, requested_height
        elif requested_width is not None:
            width = requested_width
            height = original_height * width / original_width
        elif requested_height is not None:
            height = requested_height
            width = original_width * height / original_height
        else:
            width, height = float(original_width), float(original_height)

        margin = self._image_dimension(self.state.image_m) or 0
        available_width = max(1.0, self.cell_width - 2 * margin - 8)
        available_height = max(1.0, self.cell_height - 2 * margin - 8)
        ratio = min(
            1.0,
            available_width / max(1.0, width),
            available_height / max(1.0, height),
        )
        size = (
            max(1, int(round(width * ratio))),
            max(1, int(round(height * ratio))),
        )
        resized = image.resize(size, Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(resized, master=self)

    def refresh(self):

        text = self.state.text
        photo = None
        margin = 0
        if self.state.image:
            image_name = self.state.image.replace("\\", "/").rsplit("/", 1)[-1]
            status, value = (
                self.image_provider(self.state.image)
                if self.image_provider
                else ("error", None)
            )
            if status == "ready":
                try:
                    photo = self._build_preview_photo(value)
                except Exception:
                    photo = None
            if photo is None:
                placeholder = (
                    "Loading image…"
                    if status == "loading"
                    else f"Image unavailable: {unquote(image_name)}"
                )
                text = f"{text}\n{placeholder}" if text else placeholder
            margin = int(self._image_dimension(self.state.image_m) or 0)

        self.preview_photo = photo

        anchor = {
            "Left": "w",
            "Center": "center",
            "Right": "e",
        }.get(self.state.align, "w")

        styles = []
        if self.state.bold:
            styles.append("bold")
        if self.state.italic:
            styles.append("italic")
        if self.state.underline:
            styles.append("underline")
        if self.state.strike:
            styles.append("overstrike")

        font_size = 9
        raw_font_size = self.state.font_size.strip().lower()
        for suffix in ("px", "pt"):
            if raw_font_size.endswith(suffix):
                raw_font_size = raw_font_size[:-len(suffix)].strip()
                break
        try:
            font_size = max(1, int(float(raw_font_size)))
        except ValueError:
            pass

        self.label.configure(
            text=text,
            image=photo or "",
            compound="top" if photo and text else "center",
            bg=self.state.bg,
            fg=self.state.fg,
            anchor=anchor,
            padx=max(2, margin) if self.state.image else 5,
            pady=max(1, margin) if self.state.image else 0,
            font=(
                getattr(
                    self.winfo_toplevel(),
                    "ui_font_family",
                    "Segoe UI"
                ),
                font_size,
                " ".join(styles) or "normal"
            )
        )
        self.configure(background=self.state.bg)
        self.refresh_borders()

    def refresh_borders(self):

        thickness = max(1, self.preview_border_width)
        style = str(self.preview_border_style).lower()
        no_border = self.preview_border_width <= 0 or style.endswith("-0")
        show_inner_grid = style.endswith("-1")
        dash = (6, 4) if style.startswith("dashed") else None
        if style.startswith("dotted"):
            dash = (2, 3)
        display_thickness = max(2, thickness) if dash else thickness

        placements = {
            "top": {
                "x": 0,
                "y": 0,
                "relwidth": 1,
                "height": display_thickness
            },
            "right": {
                "relx": 1,
                "x": -display_thickness,
                "y": 0,
                "width": display_thickness,
                "relheight": 1
            },
            "bottom": {
                "x": 0,
                "rely": 1,
                "y": -display_thickness,
                "relwidth": 1,
                "height": display_thickness
            },
            "left": {
                "x": 0,
                "y": 0,
                "width": display_thickness,
                "relheight": 1
            },
        }
        hidden = {
            "top": self.state.hide_top,
            "right": self.state.hide_right,
            "bottom": self.state.hide_bottom,
            "left": self.state.hide_left,
        }

        for name, edge in self.edges.items():
            duplicate_inner_edge = (
                show_inner_grid
                and name in {"right", "bottom"}
                and name not in self.outer_edges
            )
            if (
                hidden[name]
                or no_border
                or duplicate_inner_edge
                or (not show_inner_grid and name not in self.outer_edges)
            ):
                edge.place_forget()
            else:
                edge.place(**placements[name])
                edge.delete("all")
                if dash is None:
                    edge.configure(background=self.preview_border_color)
                    self.tk.call("raise", edge._w)
                    continue

                edge.configure(background=self.state.bg)
                if name in {"top", "bottom"}:
                    coordinates = (
                        0,
                        display_thickness / 2,
                        self.cell_width,
                        display_thickness / 2,
                    )
                else:
                    coordinates = (
                        display_thickness / 2,
                        0,
                        display_thickness / 2,
                        self.cell_height,
                    )
                edge.create_line(
                    *coordinates,
                    fill=self.preview_border_color,
                    width=thickness,
                    dash=dash
                )
                self.tk.call("raise", edge._w)


# ============================================================
# Designer
# ============================================================

class Designer(ttk.LabelFrame):

    def __init__(self, master):

        super().__init__(master, text="Live Preview")

        self.rows = 8
        self.cols = 5
        self.preview_cell_width = Cell.CELL_WIDTH
        self.preview_cell_height = Cell.CELL_HEIGHT
        self.preview_border_width = 1
        self.preview_border_color = "#cbd5e1"
        self.preview_border_style = "solid-1"
        self.row_scales = {}
        self.col_scales = {}

        self.cells = {}
        self.states = {}

        self.current_cell = None
        self.selection_callback = None
        self.before_selection_callback = None
        self.edit_callback = None
        self.image_previews = ImagePreviewCache(self, self._image_ready)

        self.canvas = tk.Canvas(self, highlightthickness=0, background="#f3f3f3")
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.h_scroll = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(
            yscrollcommand=self.v_scroll.set,
            xscrollcommand=self.h_scroll.set
        )

        self.canvas.grid(row=0, column=0, sticky="nsew")
        self.v_scroll.grid(row=0, column=1, sticky="ns")
        self.h_scroll.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self.grid_frame = ttk.Frame(self.canvas, padding=10)
        self.canvas_window = self.canvas.create_window(
            (0, 0),
            window=self.grid_frame,
            anchor="nw"
        )
        self.grid_frame.bind("<Configure>", self._update_scrollregion)
        self.canvas.bind(
            "<MouseWheel>",
            lambda event: self.canvas.yview_scroll(int(-event.delta / 120), "units")
        )

        self.build_grid()
        self._update_preview_title()

    def _update_scrollregion(self, event=None):

        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    @staticmethod
    def column_name(index):

        name = ""
        index += 1

        while index:
            index, remainder = divmod(index - 1, 26)
            name = chr(65 + remainder) + name

        return name

    def cell_name(self, row, col):

        return f"{self.column_name(col)}{row + 1}"

    def set_selection_callback(self, callback):

        self.selection_callback = callback

    def set_before_selection_callback(self, callback):

        self.before_selection_callback = callback

    def set_edit_callback(self, callback):

        self.edit_callback = callback

    def _cell_text_edited(self, cell):

        if self.selection_callback and self.current_cell is cell:
            self.selection_callback(cell)
        if self.edit_callback:
            self.edit_callback(cell)

    def build_grid(self):

        for w in self.grid_frame.winfo_children():
            w.destroy()

        self.cells.clear()
        self.current_cell = None

        valid_states = {}
        for r in range(self.rows):
            for c in range(self.cols):
                valid_states[(r, c)] = self.states.get((r, c), CellState())
        self.states = valid_states
        self.normalize_borders()

        ttk.Label(
            self.grid_frame,
            text=""
        ).grid(row=0, column=0)

        for c in range(self.cols):

            ttk.Label(
                self.grid_frame,
                text=self.column_name(c),
                anchor="center",
                width=14
            ).grid(
                row=0,
                column=c + 1
            )

        for r in range(self.rows):

            ttk.Label(
                self.grid_frame,
                text=str(r + 1),
                width=4
            ).grid(
                row=r + 1,
                column=0
            )

            for c in range(self.cols):

                cell = Cell(
                    self.grid_frame,
                    r,
                    c,
                    self.states[(r, c)],
                    edit_callback=self._cell_text_edited,
                    cell_width=(
                        self.preview_cell_width * self.col_scales.get(c, 1)
                    ),
                    cell_height=(
                        self.preview_cell_height * self.row_scales.get(r, 1)
                    ),
                    border_width=self.preview_border_width,
                    border_color=self.preview_border_color,
                    border_style=self.preview_border_style,
                    outer_edges={
                        name
                        for name, is_outer in (
                            ("top", r == 0),
                            ("right", c == self.cols - 1),
                            ("bottom", r == self.rows - 1),
                            ("left", c == 0),
                        )
                        if is_outer
                    },
                    image_provider=self.image_previews.get
                )

                cell.grid(
                    row=r + 1,
                    column=c + 1,
                    padx=0,
                    pady=0,
                    sticky="nsew"
                )

                cell.bind(
                    "<<CellSelected>>",
                    lambda e, rr=r, cc=c: self.select(rr, cc)
                )

                self.cells[(r, c)] = cell

        self.refresh_grid_borders()

    def select(self, row, col):

        if (
            self.current_cell
            and (self.current_cell.row, self.current_cell.col) != (row, col)
            and self.before_selection_callback
            and not self.before_selection_callback()
        ):
            return False

        if self.current_cell:

            self.current_cell.deselect()

        self.current_cell = self.cells[(row, col)]

        self.current_cell.select()

        if self.selection_callback:
            self.selection_callback(self.current_cell)

        return True

    def _image_ready(self, source):

        for cell in self.cells.values():
            if cell.state.image == source:
                cell.refresh()

    def resize(self, rows, cols):

        if rows < 1 or cols < 1:
            raise ValueError("Rows and columns must be positive")

        self.rows = rows
        self.cols = cols

        self.build_grid()
        self._update_preview_title()

    def _update_preview_title(self):

        parts = []
        if self.row_scales:
            parts.append(
                "rows " + ", ".join(
                    f"{row + 1}×{value}"
                    for row, value in sorted(self.row_scales.items())
                )
            )
        if self.col_scales:
            parts.append(
                "cols " + ", ".join(
                    f"{self.column_name(col)}×{value}"
                    for col, value in sorted(self.col_scales.items())
                )
            )
        scale = f" — {'; '.join(parts)}" if parts else ""
        self.configure(text=f"Live Preview — {self.rows} × {self.cols}{scale}")

    def set_preview_settings(
        self,
        cell_height,
        border_width,
        border_color,
        border_style,
        row_scales=None,
        col_scales=None
    ):

        self.preview_cell_height = max(20, int(cell_height) or Cell.CELL_HEIGHT)
        self.preview_border_width = max(0, int(border_width))
        self.preview_border_color = border_color or "#cbd5e1"
        self.preview_border_style = border_style or "solid-1"
        self.row_scales = dict(row_scales or {})
        self.col_scales = dict(col_scales or {})
        self._update_preview_title()

        for (row, col), cell in self.cells.items():
            cell.set_preview_geometry(
                self.preview_cell_width * self.col_scales.get(col, 1),
                self.preview_cell_height * self.row_scales.get(row, 1),
                self.preview_border_width,
                self.preview_border_color,
                self.preview_border_style
            )
        self._update_scrollregion()
        self.after_idle(self._update_scrollregion)

    def normalize_borders(self):

        for (row, col), state in self.states.items():
            if col == 0:
                state.hide_left = False
            if col == self.cols - 1:
                state.hide_right = False
            if row == 0:
                state.hide_top = False
            if row == self.rows - 1:
                state.hide_bottom = False

        for row in range(self.rows):
            for col in range(self.cols - 1):
                left = self.states[(row, col)]
                right = self.states[(row, col + 1)]
                hidden = left.hide_right or right.hide_left
                left.hide_right = hidden
                right.hide_left = hidden

        for row in range(self.rows - 1):
            for col in range(self.cols):
                top = self.states[(row, col)]
                bottom = self.states[(row + 1, col)]
                hidden = top.hide_bottom or bottom.hide_top
                top.hide_bottom = hidden
                bottom.hide_top = hidden

    def refresh_grid_borders(self):

        self.normalize_borders()
        for cell in self.cells.values():
            cell.refresh()

    def merge_dimensions(self, row, col):

        horizontal = 1
        cursor = col
        while cursor > 0 and self.states[(row, cursor)].hide_left:
            horizontal += 1
            cursor -= 1
        cursor = col
        while cursor < self.cols - 1 and self.states[(row, cursor)].hide_right:
            horizontal += 1
            cursor += 1

        vertical = 1
        cursor = row
        while cursor > 0 and self.states[(cursor, col)].hide_top:
            vertical += 1
            cursor -= 1
        cursor = row
        while cursor < self.rows - 1 and self.states[(cursor, col)].hide_bottom:
            vertical += 1
            cursor += 1

        return horizontal, vertical

# ============================================================
# Properties
# ============================================================

class PropertiesPanel(ttk.LabelFrame):

    def __init__(self, master, designer):

        super().__init__(master, text="Properties")

        self.designer = designer
        self.status_callback = None
        self.before_resize_callback = None
        self.after_resize_callback = None
        self.preview_callback = None

        padx = 8
        pady = 6

        ttk.Label(self, text="Rows").grid(row=0, column=0, sticky="w", padx=padx, pady=pady)

        self.rows = ttk.Spinbox(
            self,
            from_=1,
            to=100,
            width=8
        )

        self.rows.set(8)
        self.rows.grid(row=0, column=1)

        ttk.Label(self, text="Cols").grid(row=1, column=0, sticky="w", padx=padx, pady=pady)

        self.cols = ttk.Spinbox(
            self,
            from_=1,
            to=50,
            width=8
        )

        self.cols.set(5)
        self.cols.grid(row=1, column=1)

        ttk.Separator(self).grid(row=2, column=0, columnspan=2, sticky="ew", padx=padx, pady=8)

        ttk.Label(self, text="Border Width").grid(row=3, column=0, sticky="w", padx=padx, pady=pady)
        self.border_width = ttk.Spinbox(self, from_=0, to=10, width=8)
        self.border_width.set(1)
        self.border_width.grid(row=3, column=1)

        ttk.Label(self, text="Border Color").grid(row=4, column=0, sticky="w", padx=padx, pady=pady)
        self.border_color = ttk.Entry(self, width=11)
        self.border_color.insert(0, "#cbd5e1")
        self.border_color.grid(row=4, column=1)

        ttk.Label(self, text="Border Style").grid(row=5, column=0, sticky="w", padx=padx, pady=pady)
        self.border_style = ttk.Combobox(
            self,
            values=[
                "solid-1",
                "dashed-1",
                "dotted-1",
                "solid",
                "dashed",
                "dotted",
                "solid-0"
            ],
            state="readonly",
            width=9
        )
        self.border_style.set("solid-1")
        self.border_style.grid(row=5, column=1)

        ttk.Label(self, text="Cell Height").grid(row=6, column=0, sticky="w", padx=padx, pady=pady)
        self.cell_height = ttk.Spinbox(self, from_=0, to=300, width=8)
        self.cell_height.set(36)
        self.cell_height.grid(row=6, column=1)

        self.stretch = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            self,
            text="Auto Stretch",
            variable=self.stretch,
            command=self._notify_preview
        ).grid(row=7, column=0, columnspan=2, sticky="w", padx=padx, pady=pady)

        ttk.Label(self, text="Scale").grid(
            row=8,
            column=0,
            sticky="w",
            padx=padx,
            pady=(pady, 1)
        )
        self.scale = ttk.Entry(self, width=18)
        self.scale.grid(
            row=9,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=padx
        )
        ttk.Label(
            self,
            text="Example: A1=2,3;C2=1,2",
            foreground="#6b7280"
        ).grid(
            row=10,
            column=0,
            columnspan=2,
            sticky="w",
            padx=padx,
            pady=(1, 4)
        )

        ttk.Button(
            self,
            text="Apply Grid",
            command=self.apply
        ).grid(
            row=11,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=padx,
            pady=20
        )

        for widget in (
            self.border_width,
            self.border_color,
            self.border_style,
            self.cell_height,
            self.scale,
        ):
            widget.bind("<KeyRelease>", self._notify_preview)
            widget.bind("<FocusOut>", self._notify_preview)
        self.border_style.bind("<<ComboboxSelected>>", self._notify_preview)
        self.border_width.configure(command=self._notify_preview)
        self.cell_height.configure(command=self._notify_preview)

    def set_status_callback(self, callback):

        self.status_callback = callback

    def set_before_resize_callback(self, callback):

        self.before_resize_callback = callback

    def set_after_resize_callback(self, callback):

        self.after_resize_callback = callback

    def set_preview_callback(self, callback):

        self.preview_callback = callback

    def _notify_preview(self, event=None):

        if not self.preview_callback:
            return
        try:
            values = self.get_values()
            self.winfo_rgb(values["border_color"])
        except (ValueError, tk.TclError):
            return
        self.preview_callback(values)

    def _status(self, text):

        if self.status_callback:
            self.status_callback(text)

    def set_dimensions(self, rows, cols):

        self.rows.delete(0, "end")
        self.rows.insert(0, str(rows))
        self.cols.delete(0, "end")
        self.cols.insert(0, str(cols))

    @staticmethod
    def _replace_entry(entry, value):

        entry.delete(0, "end")
        entry.insert(0, str(value))

    def load_model_settings(self, model, scale_text=None):

        if scale_text is None:
            coordinates = {}
            for row, value in model.row_scales.items():
                coordinates.setdefault((row, 0), [1, 1])[0] = value
            for col, value in model.col_scales.items():
                coordinates.setdefault((0, col), [1, 1])[1] = value
            scale_text = ";".join(
                f"{self.designer.column_name(col)}{row + 1}="
                f"{values[0]},{values[1]}"
                for (row, col), values in sorted(coordinates.items())
            )
        self.load_values({
            "rows": model.rows,
            "cols": model.cols,
            "border_width": model.border_width,
            "border_color": model.border_color,
            "border_style": model.border_style,
            "stretch": model.stretch,
            "cell_height": model.cell_height,
            "scale": scale_text,
        })
        self._notify_preview()

    def load_values(self, values):

        self.set_dimensions(values["rows"], values["cols"])
        self._replace_entry(self.border_width, values["border_width"])
        self._replace_entry(self.border_color, values["border_color"])
        self.border_style.set(values["border_style"])
        self._replace_entry(self.cell_height, values["cell_height"])
        self.stretch.set(bool(values["stretch"]))
        self._replace_entry(self.scale, values.get("scale", ""))

    def get_values(self):

        rows = int(self.rows.get())
        cols = int(self.cols.get())
        border_width = int(self.border_width.get())
        cell_height = int(self.cell_height.get())
        border_color = self.border_color.get().strip()

        if not 1 <= rows <= 100:
            raise ValueError("Rows must be between 1 and 100")
        if not 1 <= cols <= 50:
            raise ValueError("Columns must be between 1 and 50")
        if not 0 <= border_width <= 10:
            raise ValueError("Border width must be between 0 and 10")
        if not 0 <= cell_height <= 300:
            raise ValueError("Cell height must be between 0 and 300")
        if not border_color:
            raise ValueError("Border color cannot be empty")

        border_style = self.border_style.get()
        scale = self.scale.get().strip()
        color = border_color.replace('"', "")
        preamble = (
            f'{rows},{cols},{border_width},"{color}","{border_style}",'
            f'{int(self.stretch.get())},{cell_height}'
        )
        if scale:
            preamble += f",scale({scale})"
        model = tentags.compile(preamble, "style()", "data()")

        return {
            "rows": rows,
            "cols": cols,
            "border_width": border_width,
            "border_color": border_color,
            "border_style": border_style,
            "stretch": int(self.stretch.get()),
            "cell_height": cell_height,
            "scale": scale,
            "row_scales": model.row_scales,
            "col_scales": model.col_scales,
            "preamble": preamble,
        }

    def get_preamble(self):

        values = self.get_values()
        return values["preamble"]

    def apply(self):

        try:
            values = self.get_values()
            if (
                self.before_resize_callback
                and not self.before_resize_callback(values)
            ):
                return
            self.load_values(values)
            self.designer.resize(values["rows"], values["cols"])
            if self.preview_callback:
                self.preview_callback(values)
            if self.after_resize_callback:
                self.after_resize_callback(values)
            self._status(
                f'Grid and lower editor resized to '
                f'{values["rows"]} × {values["cols"]}'
            )
        except Exception as exc:
            messagebox.showerror("Invalid grid settings", str(exc), parent=self)


# ============================================================
# Cell Properties
# ============================================================

class CellProperties(ttk.LabelFrame):

    def __init__(self, master, designer):

        super().__init__(master, text="Cell Properties")

        self.designer = designer
        self.mode = "cell"
        self.status_callback = None
        self.style_change_callback = None
        self.data_change_callback = None
        self.output_change_callback = None
        self.build()
        self.designer.set_before_selection_callback(self.apply_before_selection)
        self.designer.set_selection_callback(self.load_cell)

    def build(self):

        pad = 8

        menu = ttk.Frame(self)
        menu.pack(fill="x", padx=4, pady=(6, 4))
        self.mode_buttons = {}
        menu_items = (
            ("cell", "Cell"),
            ("range", "Row / Col"),
            ("table", "Table"),
            ("output", "Output"),
        )
        for column, (mode, label) in enumerate(menu_items):
            button = ttk.Button(
                menu,
                text=label,
                command=lambda selected=mode: self.show_mode(selected)
            )
            button.grid(row=0, column=column, sticky="ew", padx=1)
            menu.columnconfigure(column, weight=1)
            self.mode_buttons[mode] = button

        self.editor_panel = ttk.Frame(self)
        self.output_panel = ttk.Frame(self)

        self.range_panel = ttk.LabelFrame(
            self.editor_panel,
            text="Apply style to"
        )
        self.range_target = tk.StringVar(value="Row")
        ttk.Radiobutton(
            self.range_panel,
            text="Selected row",
            variable=self.range_target,
            value="Row"
        ).pack(anchor="w", padx=pad, pady=(5, 1))
        ttk.Radiobutton(
            self.range_panel,
            text="Selected column",
            variable=self.range_target,
            value="Column"
        ).pack(anchor="w", padx=pad, pady=(1, 5))

        self.style_panel = ttk.LabelFrame(
            self.editor_panel,
            text="Style"
        )

        row = ttk.Frame(self.style_panel)
        row.pack(fill="x", padx=pad, pady=(7, 2))

        self.bold = tk.BooleanVar()
        ttk.Checkbutton(
            row,
            text="Bold",
            variable=self.bold
        ).pack(side="left")

        self.italic = tk.BooleanVar()
        ttk.Checkbutton(
            row,
            text="Italic",
            variable=self.italic
        ).pack(side="left", padx=(12, 0))

        row = ttk.Frame(self.style_panel)
        row.pack(fill="x", padx=pad, pady=(2, 7))

        self.underline = tk.BooleanVar()
        ttk.Checkbutton(
            row,
            text="Underline",
            variable=self.underline
        ).pack(side="left")

        self.strike = tk.BooleanVar()
        ttk.Checkbutton(
            row,
            text="Strike",
            variable=self.strike
        ).pack(side="left", padx=(12, 0))

        ttk.Label(
            self.style_panel,
            text="Font Size (<fs>)"
        ).pack(anchor="w", padx=pad)

        self.font_size = ttk.Spinbox(
            self.style_panel,
            from_=1,
            to=99,
            increment=1,
            width=5
        )
        self.font_size.pack(anchor="w", padx=pad, pady=(2, 6))

        ttk.Label(
            self.style_panel,
            text="Alignment"
        ).pack(anchor="w", padx=pad)

        self.align = ttk.Combobox(
            self.style_panel,
            values=[
                "Left",
                "Center",
                "Right"
            ],
            state="readonly"
        )

        self.align.current(0)
        self.align.pack(fill="x", padx=pad, pady=4)

        ttk.Label(
            self.style_panel,
            text="Text Color"
        ).pack(anchor="w", padx=pad)

        row = ttk.Frame(self.style_panel)
        row.pack(fill="x", padx=pad)

        self.text_color = "#000000"

        self.text_preview = tk.Label(
            row,
            bg=self.text_color,
            width=4,
            relief="solid"
        )

        self.text_preview.pack(side="left")

        ttk.Button(
            row,
            text="Choose...",
            command=self.choose_text_color
        ).pack(side="left", padx=8)

        ttk.Label(
            self.style_panel,
            text="Background"
        ).pack(anchor="w", padx=pad, pady=(10, 0))

        row = ttk.Frame(self.style_panel)
        row.pack(fill="x", padx=pad, pady=(0, 8))

        self.bg_color = "#ffffff"

        self.bg_preview = tk.Label(
            row,
            bg=self.bg_color,
            width=4,
            relief="solid"
        )

        self.bg_preview.pack(side="left")

        ttk.Button(
            row,
            text="Choose...",
            command=self.choose_bg_color
        ).pack(side="left", padx=8)

        self.content_panel = ttk.LabelFrame(
            self.editor_panel,
            text="Cell content"
        )

        ttk.Label(
            self.content_panel,
            text="Text"
        ).pack(anchor="w", padx=pad, pady=(7, 2))

        self.text = tk.Text(
            self.content_panel,
            height=3,
            width=28
        )
        self.text.pack(fill="x", padx=pad)

        ttk.Label(
            self.content_panel,
            text="URL"
        ).pack(anchor="w", padx=pad, pady=(7, 2))

        self.url = ttk.Entry(self.content_panel)
        self.url.pack(fill="x", padx=pad)

        ttk.Label(
            self.content_panel,
            text="Mark"
        ).pack(anchor="w", padx=pad, pady=(7, 2))

        self.mark = ttk.Entry(self.content_panel)
        self.mark.pack(fill="x", padx=pad)

        ttk.Label(
            self.content_panel,
            text="Image"
        ).pack(anchor="w", padx=pad, pady=(7, 2))

        row = ttk.Frame(self.content_panel)
        row.pack(fill="x", padx=pad)

        self.image = ttk.Entry(row)

        self.image.pack(
            side="left",
            fill="x",
            expand=True
        )

        ttk.Button(
            row,
            text="Browse...",
            command=self.select_image
        ).pack(side="left", padx=6)

        row = ttk.Frame(self.content_panel)
        row.pack(fill="x", padx=pad, pady=(7, 8))

        ttk.Label(row, text="W").grid(row=0, column=0)

        self.w = ttk.Entry(row, width=6)
        self.w.insert(0, "120")
        self.w.grid(row=0, column=1)

        ttk.Label(row, text="H").grid(row=0, column=2, padx=(10, 0))

        self.h = ttk.Entry(row, width=6)
        self.h.insert(0, "auto")
        self.h.grid(row=0, column=3)

        ttk.Label(row, text="M").grid(row=0, column=4, padx=(10, 0))

        self.m = ttk.Entry(row, width=6)
        self.m.insert(0, "15")
        self.m.grid(row=0, column=5)

        self.apply_button = ttk.Button(
            self.editor_panel,
            text="Apply",
            command=self.apply_from_button
        )

        self._build_output_panel(pad)
        self.show_mode("cell", save_current=False)

    def _build_output_panel(self, pad):

        formats = ttk.LabelFrame(self.output_panel, text="Formats")
        formats.pack(fill="x", padx=pad, pady=(5, 4))

        self.var_html = tk.BooleanVar(value=True)
        self.var_pdf = tk.BooleanVar(value=True)
        self.var_xlsx = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            formats,
            text="HTML",
            variable=self.var_html
        ).pack(side="left", padx=6, pady=5)
        ttk.Checkbutton(
            formats,
            text="PDF",
            variable=self.var_pdf
        ).pack(side="left", padx=6, pady=5)
        ttk.Checkbutton(
            formats,
            text="XLSX",
            variable=self.var_xlsx
        ).pack(side="left", padx=6, pady=5)

        html_box = ttk.LabelFrame(self.output_panel, text="HTML")
        html_box.pack(fill="x", padx=pad, pady=4)
        self.html_full_page = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            html_box,
            text="Full HTML document",
            variable=self.html_full_page
        ).pack(anchor="w", padx=6, pady=(5, 2))
        ttk.Label(html_box, text="Page title").pack(anchor="w", padx=6)
        self.html_title = ttk.Entry(html_box)
        self.html_title.insert(0, "TenTags Report")
        self.html_title.pack(fill="x", padx=6, pady=(2, 6))

        pdf_box = ttk.LabelFrame(self.output_panel, text="PDF")
        pdf_box.pack(fill="x", padx=pad, pady=4)
        row = ttk.Frame(pdf_box)
        row.pack(fill="x", padx=6, pady=(5, 2))
        ttk.Label(row, text="Page").pack(side="left")
        self.pdf_page_size = ttk.Combobox(
            row,
            values=["A3", "A4", "A5", "letter", "legal", "tabloid"],
            state="readonly",
            width=8
        )
        self.pdf_page_size.set("A4")
        self.pdf_page_size.pack(side="left", padx=(4, 8))
        ttk.Label(row, text="Orientation").pack(side="left")
        self.pdf_orientation = ttk.Combobox(
            row,
            values=["portrait", "landscape"],
            state="readonly",
            width=10
        )
        self.pdf_orientation.set("portrait")
        self.pdf_orientation.pack(side="left", padx=4)

        ttk.Label(
            pdf_box,
            text="Margins: left / right / top / bottom (pt)"
        ).pack(anchor="w", padx=6, pady=(3, 1))
        row = ttk.Frame(pdf_box)
        row.pack(fill="x", padx=6, pady=(0, 6))
        self.pdf_margins = []
        for _ in range(4):
            entry = ttk.Spinbox(row, from_=0, to=500, width=5)
            entry.set("36")
            entry.pack(side="left", padx=(0, 4))
            self.pdf_margins.append(entry)

        xlsx_box = ttk.LabelFrame(self.output_panel, text="XLSX")
        xlsx_box.pack(fill="x", padx=pad, pady=4)
        row = ttk.Frame(xlsx_box)
        row.pack(fill="x", padx=6, pady=(5, 2))
        ttk.Label(row, text="Paper").pack(side="left")
        self.xlsx_page_size = ttk.Combobox(
            row,
            values=["A3", "A4", "A5", "letter", "legal", "tabloid"],
            state="readonly",
            width=8
        )
        self.xlsx_page_size.set("A4")
        self.xlsx_page_size.pack(side="left", padx=(4, 8))
        ttk.Label(row, text="Orientation").pack(side="left")
        self.xlsx_orientation = ttk.Combobox(
            row,
            values=["portrait", "landscape"],
            state="readonly",
            width=10
        )
        self.xlsx_orientation.set("portrait")
        self.xlsx_orientation.pack(side="left", padx=4)

        ttk.Label(xlsx_box, text="Sheet name").pack(anchor="w", padx=6)
        self.xlsx_sheet_name = ttk.Entry(xlsx_box)
        self.xlsx_sheet_name.insert(0, "Table")
        self.xlsx_sheet_name.pack(fill="x", padx=6, pady=(2, 4))
        row = ttk.Frame(xlsx_box)
        row.pack(fill="x", padx=6, pady=(0, 5))
        self.xlsx_fit_to_page = tk.BooleanVar(value=True)
        self.xlsx_gridlines = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            row,
            text="Fit to page",
            variable=self.xlsx_fit_to_page
        ).pack(side="left")
        ttk.Checkbutton(
            row,
            text="Gridlines",
            variable=self.xlsx_gridlines
        ).pack(side="left", padx=(10, 0))

        ttk.Button(
            self.output_panel,
            text="Apply Output Settings",
            command=self.apply_output_settings
        ).pack(fill="x", padx=pad, pady=(6, 10))

    def show_mode(self, mode, save_current=True):

        pad = 8
        if mode not in self.mode_buttons:
            raise ValueError(f"Unknown properties mode: {mode}")
        if save_current and self.mode == "cell" and mode != "cell":
            if not self.apply():
                return

        self.mode = mode
        for name, button in self.mode_buttons.items():
            button.configure(state="disabled" if name == mode else "normal")

        self.editor_panel.pack_forget()
        self.output_panel.pack_forget()
        self.range_panel.pack_forget()
        self.style_panel.pack_forget()
        self.content_panel.pack_forget()
        self.apply_button.pack_forget()

        if mode == "output":
            self.configure(text="Output Settings")
            self.output_panel.pack(fill="both", expand=True)
            self._status("Output settings")
            return

        self.editor_panel.pack(fill="both", expand=True)
        if mode == "range":
            self.range_panel.pack(fill="x", padx=pad, pady=(5, 2))
        self.style_panel.pack(fill="x", padx=pad, pady=(5, 2))
        if mode == "cell":
            self.content_panel.pack(fill="x", padx=pad, pady=(2, 2))
        self.apply_button.pack(fill="x", padx=pad, pady=(6, 10))

        cell = self.designer.current_cell
        if cell:
            self.load_cell(cell)

    def get_output_settings(self):

        margins = tuple(float(entry.get()) for entry in self.pdf_margins)
        if any(value < 0 for value in margins):
            raise ValueError("PDF margins cannot be negative")

        sheet_name = self.xlsx_sheet_name.get().strip() or "Table"
        if len(sheet_name) > 31 or any(char in sheet_name for char in '[]:*?/\\'):
            raise ValueError(
                "XLSX sheet name must be at most 31 characters and cannot "
                "contain []:*?/\\"
            )

        return {
            "html": self.var_html.get(),
            "pdf": self.var_pdf.get(),
            "xlsx": self.var_xlsx.get(),
            "html_full_page": self.html_full_page.get(),
            "html_title": self.html_title.get().strip() or "TenTags Report",
            "pdf_page_size": self.pdf_page_size.get(),
            "pdf_orientation": self.pdf_orientation.get(),
            "pdf_margins": margins,
            "xlsx_page_size": self.xlsx_page_size.get(),
            "xlsx_orientation": self.xlsx_orientation.get(),
            "xlsx_sheet_name": sheet_name,
            "xlsx_fit_to_page": self.xlsx_fit_to_page.get(),
            "xlsx_gridlines": self.xlsx_gridlines.get(),
        }

    def choose_text_color(self):

        color = colorchooser.askcolor(
            initialcolor=self.text_color,
            parent=self
        )[1]

        if color:
            self.text_color = color
            self.text_preview.configure(bg=color)

    def choose_bg_color(self):

        color = colorchooser.askcolor(
            initialcolor=self.bg_color,
            parent=self
        )[1]

        if color:
            self.bg_color = color
            self.bg_preview.configure(bg=color)

    def select_image(self):

        filename = filedialog.askopenfilename(
            parent=self,
            filetypes=[
                ("Images", "*.png *.jpg *.jpeg *.gif *.bmp")
            ]
        )

        if filename:
            self.image.delete(0, "end")
            self.image.insert(0, filename)

    def set_status_callback(self, callback):

        self.status_callback = callback

    def set_style_change_callback(self, callback):

        self.style_change_callback = callback

    def set_data_change_callback(self, callback):

        self.data_change_callback = callback

    def set_output_change_callback(self, callback):

        self.output_change_callback = callback

    def apply_output_settings(self):

        try:
            self.get_output_settings()
            if self.output_change_callback:
                self.output_change_callback()
        except Exception as exc:
            messagebox.showerror("Invalid output settings", str(exc), parent=self)
            self._status("Output settings were not applied")
            return False

        self._status("Output settings synchronized — click Regenerate")
        return True

    def _status(self, text):

        if self.status_callback:
            self.status_callback(text)

    @staticmethod
    def _replace_entry(entry, value):

        entry.delete(0, "end")
        entry.insert(0, value)

    def load_cell(self, cell):

        state = cell.state
        address = self.designer.cell_name(cell.row, cell.col)
        titles = {
            "cell": f"Cell {address}",
            "range": f"Row / Column from {address}",
            "table": "Table Style",
            "output": "Output Settings",
        }
        self.configure(text=titles[self.mode])

        self.text.delete("1.0", "end")
        self.text.insert("1.0", state.text)
        self.bold.set(state.bold)
        self.italic.set(state.italic)
        self.underline.set(state.underline)
        self.strike.set(state.strike)
        self._replace_entry(self.font_size, state.font_size)
        self.align.set(state.align)

        self.text_color = state.fg
        self.bg_color = state.bg
        self.text_preview.configure(bg=state.fg)
        self.bg_preview.configure(bg=state.bg)

        self._replace_entry(self.url, state.url)
        self._replace_entry(self.mark, state.mark)
        self._replace_entry(self.image, state.image)
        self._replace_entry(self.w, state.image_w)
        self._replace_entry(self.h, state.image_h)
        self._replace_entry(self.m, state.image_m)

        horizontal, vertical = self.designer.merge_dimensions(cell.row, cell.col)
        merge_parts = []
        if horizontal > 1:
            merge_parts.append(f"cm:{horizontal}")
        if vertical > 1:
            merge_parts.append(f"rm:{vertical}")
        merge = f" — {'; '.join(merge_parts)}" if merge_parts else ""
        self._status(f"Selected {address}{merge}")

    def apply_before_selection(self):

        if self.mode == "cell":
            return self.apply()
        return True

    def apply_from_button(self):

        return self.apply(force_style_sync=True)

    def _target_cells(self, cell):

        if self.mode == "cell":
            return [cell]
        if self.mode == "range" and self.range_target.get() == "Row":
            return [
                self.designer.cells[(cell.row, col)]
                for col in range(self.designer.cols)
            ]
        if self.mode == "range":
            return [
                self.designer.cells[(row, cell.col)]
                for row in range(self.designer.rows)
            ]
        if self.mode == "table":
            return [
                self.designer.cells[(row, col)]
                for row in range(self.designer.rows)
                for col in range(self.designer.cols)
            ]
        return [cell]

    @staticmethod
    def _style_signature(state):

        return (
            state.bold,
            state.italic,
            state.underline,
            state.strike,
            state.font_size,
            state.align,
            state.fg,
            state.bg,
        )

    def _apply_style_to_state(self, state):

        state.bold = self.bold.get()
        state.italic = self.italic.get()
        state.underline = self.underline.get()
        state.strike = self.strike.get()
        state.font_size = self.font_size.get().strip()
        state.align = self.align.get()
        state.fg = self.text_color
        state.bg = self.bg_color

    def apply(self, force_style_sync=False):

        cell = self.designer.current_cell
        if not cell:
            messagebox.showinfo("Cell properties", "Select a cell first", parent=self)
            return False

        try:
            self.winfo_rgb(self.text_color)
            self.winfo_rgb(self.bg_color)
        except tk.TclError:
            messagebox.showerror("Invalid color", "Choose valid text and background colors", parent=self)
            return False

        target_cells = self._target_cells(cell)
        previous_styles = [
            self._style_signature(target.state)
            for target in target_cells
        ]

        state = cell.state
        previous_data = (
            state.text,
            state.url,
            state.mark,
            state.image,
            state.image_w,
            state.image_h,
            state.image_m,
        )

        for target in target_cells:
            self._apply_style_to_state(target.state)

        if self.mode == "cell":
            state.text = self.text.get("1.0", "end-1c")
            state.url = self.url.get().strip()
            state.mark = self.mark.get().strip()
            state.image = self.image.get().strip()
            state.image_w = self.w.get().strip() or "120"
            state.image_h = self.h.get().strip() or "auto"
            state.image_m = self.m.get().strip() or "0"

        current_styles = [
            self._style_signature(target.state)
            for target in target_cells
        ]
        if (
            (previous_styles != current_styles or force_style_sync)
            and self.style_change_callback
        ):
            self.style_change_callback()

        current_data = (
            state.text,
            state.url,
            state.mark,
            state.image,
            state.image_w,
            state.image_h,
            state.image_m,
        )
        if previous_data != current_data and self.data_change_callback:
            self.data_change_callback()

        for target in target_cells:
            target.refresh()

        address = self.designer.cell_name(cell.row, cell.col)
        if self.mode == "range":
            scope = (
                f"row {cell.row + 1}"
                if self.range_target.get() == "Row"
                else f"column {self.designer.column_name(cell.col)}"
            )
        elif self.mode == "table":
            scope = "the whole table"
        else:
            scope = f"cell {address}"
        self._status(f"Applied style to {scope}")
        return True

# ============================================================
# Toolbar
# ============================================================
# Toolbar
# ============================================================

class Toolbar(ttk.Frame):

    def __init__(self, master):

        super().__init__(master)

        self.pack(fill="x")

        ttk.Label(
            self,
            text="🏷 TenTags Studio",
            font=(
                getattr(
                    self.winfo_toplevel(),
                    "ui_font_family",
                    "Segoe UI"
                ),
                13,
                "bold"
            )
        ).pack(
            side="left",
            padx=10,
            pady=8
        )

        import webbrowser

        style = ttk.Style()
        style.configure(
            "Link.TLabel",
            foreground="#2563eb",
            font=(
                getattr(
                    self.winfo_toplevel(),
                    "ui_font_family",
                    "Segoe UI"
                ),
                11,
                "underline"
            )
        )

        self.link_lbl = ttk.Label(
            self,
            text="tentags.org",
            style="Link.TLabel",
            cursor="hand2"
        )
        self.link_lbl.place(relx=0.5, rely=0.5, anchor="center")
        self.link_lbl.bind("<Button-1>", lambda e: webbrowser.open("https://tentags.org"))

        self.copy_btn = ttk.Button(
            self,
            text="Copy Code"
        )

        self.copy_btn.pack(
            side="right",
            padx=3
        )

        self.regenerate_btn = ttk.Button(
            self,
            text="Regenerate"
        )

        self.regenerate_btn.pack(
            side="right",
            padx=3,
            before=self.copy_btn
        )

        self.refresh_btn = ttk.Button(
            self,
            text="Refresh"
        )

        self.refresh_btn.pack(
            side="right",
            padx=3,
            before=self.regenerate_btn
        )

    def show_regenerate(self):

        if not self.regenerate_btn.winfo_manager():
            self.regenerate_btn.pack(
                side="right",
                padx=3,
                before=self.copy_btn
            )

    def hide_regenerate(self):

        self.regenerate_btn.configure(state="normal", text="Regenerate")


# ============================================================
# Code Editor
# ============================================================


# ============================================================
# Code Editor
# ============================================================

class CodeEditor(ttk.LabelFrame):

    INDENT = "    "

    def __init__(self, master):

        super().__init__(
            master,
            text="Generated TenTags / Python",
            padding=(15, 5)
        )

        self.modified_callback = None
        self.suppress_modified = False

        self.font = tkfont.Font(
            family=getattr(
                self.winfo_toplevel(),
                "editor_font_family",
                "Consolas"
            ),
            size=11
        )

        self.linenumbers = tk.Text(
            self,
            font=self.font,
            width=4,
            padx=5,
            pady=5,
            takefocus=0,
            border=0,
            background="#f3f3f3",
            foreground="#a1a1a1",
            state="disabled",
            wrap="none"
        )
        self.linenumbers.tag_configure("right_align", justify="right")

        self.text = tk.Text(
            self,
            font=self.font,
            undo=True,
            autoseparators=True,
            maxundo=-1,
            wrap="none",
            height=16,
            padx=10,
            pady=5
        )

        def on_scrollbar(*args):
            self.text.yview(*args)
            self.linenumbers.yview(*args)

        vs = ttk.Scrollbar(
            self,
            orient="vertical",
            command=on_scrollbar
        )

        hs = ttk.Scrollbar(
            self,
            orient="horizontal",
            command=self.text.xview
        )

        def on_text_scroll(*args):
            vs.set(*args)
            self.linenumbers.yview_moveto(args[0])

        self.text.configure(
            yscrollcommand=on_text_scroll,
            xscrollcommand=hs.set
        )

        # Editor toolbar
        self.editor_toolbar = ttk.Frame(self)
        self.editor_toolbar.grid(
            row=0,
            column=0,
            columnspan=3,
            sticky="ew",
            pady=(2, 5)
        )

        self.import_style_btn = ttk.Button(
            self.editor_toolbar,
            text="Import Style as CSV",
            command=self.import_style_csv
        )
        self.import_style_btn.pack(side="left", padx=5)

        self.import_data_btn = ttk.Button(
            self.editor_toolbar,
            text="Import Data as CSV",
            command=self.import_data_csv
        )
        self.import_data_btn.pack(side="left", padx=5)

        self.linenumbers.grid(
            row=1,
            column=0,
            sticky="ns"
        )

        self.text.grid(
            row=1,
            column=1,
            sticky="nsew"
        )

        vs.grid(
            row=1,
            column=2,
            sticky="ns"
        )

        hs.grid(
            row=2,
            column=1,
            sticky="ew"
        )

        self.rowconfigure(1, weight=1)
        self.columnconfigure(1, weight=1)

        self.linenumbers.bind(
            "<MouseWheel>",
            lambda event: self.text.yview_scroll(int(-event.delta / 120), "units")
        )

        self.text.bind("<<Modified>>", self._on_modified)
        self.text.bind("<Tab>", self._indent)
        self.text.bind("<Shift-Tab>", self._unindent)
        self.text.bind("<ISO_Left_Tab>", self._unindent)
        self.text.bind("<Return>", self._insert_newline)
        self.text.bind("<KP_Enter>", self._insert_newline)
        self.text.bind("<Control-KeyPress>", self._control_shortcut)

        self.load_template()

    def _selection(self):

        try:
            return self.text.index("sel.first"), self.text.index("sel.last")
        except tk.TclError:
            return None

    def _selected_line_numbers(self):

        selection = self._selection()
        if selection is None:
            line = int(self.text.index("insert").split(".", 1)[0])
            return line, line

        start, end = selection
        first_line = int(start.split(".", 1)[0])
        end_line, end_column = (int(part) for part in end.split(".", 1))
        if end_column == 0 and end_line > first_line:
            end_line -= 1
        return first_line, end_line

    def _select_lines(self, first_line, last_line):

        start = f"{first_line}.0"
        end = f"{last_line}.end"
        self.text.tag_remove("sel", "1.0", "end")
        self.text.tag_add("sel", start, end)
        self.text.mark_set("insert", end)
        self.text.see("insert")

    def _indent(self, event=None):

        selection = self._selection()
        if selection is None:
            column = int(self.text.index("insert").split(".", 1)[1])
            width = len(self.INDENT) - (column % len(self.INDENT))
            self.text.insert("insert", " " * width)
            return "break"

        first_line, last_line = self._selected_line_numbers()
        self.text.edit_separator()
        for line in range(first_line, last_line + 1):
            self.text.insert(f"{line}.0", self.INDENT)
        self.text.edit_separator()
        self._select_lines(first_line, last_line)
        return "break"

    def _unindent(self, event=None):

        had_selection = self._selection() is not None
        first_line, last_line = self._selected_line_numbers()
        insert_line, insert_column = (
            int(part) for part in self.text.index("insert").split(".", 1)
        )
        removed_from_insert_line = 0

        self.text.edit_separator()
        for line in range(first_line, last_line + 1):
            start = f"{line}.0"
            prefix = self.text.get(start, f"{line}.0+{len(self.INDENT)}c")
            if prefix.startswith("\t"):
                remove_count = 1
            else:
                remove_count = len(prefix) - len(prefix.lstrip(" "))
                remove_count = min(remove_count, len(self.INDENT))
            if remove_count:
                self.text.delete(start, f"{start}+{remove_count}c")
                if line == insert_line:
                    removed_from_insert_line = remove_count
        self.text.edit_separator()

        if had_selection:
            self._select_lines(first_line, last_line)
        else:
            new_column = max(0, insert_column - removed_from_insert_line)
            self.text.mark_set("insert", f"{insert_line}.{new_column}")
        return "break"

    def _insert_newline(self, event=None):

        selection = self._selection()
        if selection is not None:
            self.text.delete(*selection)

        before_cursor = self.text.get("insert linestart", "insert")
        indent = before_cursor[:len(before_cursor) - len(before_cursor.lstrip(" \t"))]
        self.text.insert("insert", "\n" + indent)
        self.text.see("insert")
        return "break"

    def _control_shortcut(self, event):

        keysym = str(getattr(event, "keysym", "")).lower()
        keycode = getattr(event, "keycode", None)
        character = getattr(event, "char", "")
        shift_pressed = bool(getattr(event, "state", 0) & 0x1)

        shortcuts = {
            "a": self._select_all,
            "c": self._copy,
            "v": self._paste,
            "x": self._cut,
            "y": self._redo,
            "z": self._redo if shift_pressed else self._undo,
        }
        control_characters = {
            "\x01": "a",
            "\x03": "c",
            "\x16": "v",
            "\x18": "x",
            "\x19": "y",
            "\x1a": "z",
        }
        windows_keycodes = {
            65: "a",
            67: "c",
            86: "v",
            88: "x",
            89: "y",
            90: "z",
        }
        shortcut = (
            keysym
            if keysym in shortcuts
            else control_characters.get(character)
            or windows_keycodes.get(keycode)
        )
        handler = shortcuts.get(shortcut)
        if handler:
            return handler(event)
        return None

    def _select_all(self, event=None):

        self.text.tag_add("sel", "1.0", "end-1c")
        self.text.mark_set("insert", "end-1c")
        self.text.see("insert")
        return "break"

    def _copy(self, event=None):

        self.text.event_generate("<<Copy>>")
        return "break"

    def _cut(self, event=None):

        self.text.event_generate("<<Cut>>")
        return "break"

    def _paste(self, event=None):

        self.text.event_generate("<<Paste>>")
        return "break"

    def _undo(self, event=None):

        try:
            self.text.edit_undo()
        except tk.TclError:
            pass
        return "break"

    def _redo(self, event=None):

        try:
            self.text.edit_redo()
        except tk.TclError:
            pass
        return "break"

    def set_modified_callback(self, callback):

        self.modified_callback = callback

    def _on_modified(self, event=None):

        if self.suppress_modified:
            self.text.edit_modified(False)
            return

        if self.text.edit_modified():
            self.text.edit_modified(False)
            self.update_line_numbers()
            if self.modified_callback:
                self.modified_callback()

    def update_line_numbers(self):

        if not hasattr(self, "linenumbers"):
            return

        self.linenumbers.configure(state="normal")
        self.linenumbers.delete("1.0", "end")

        line_count = int(self.text.index("end-1c").split(".")[0])
        numbers_str = "\n".join(str(i) for i in range(1, line_count + 1))

        width = max(3, len(str(line_count)))
        self.linenumbers.configure(width=width)

        self.linenumbers.insert("1.0", numbers_str, "right_align")
        self.linenumbers.configure(state="disabled")
        self.linenumbers.yview_moveto(self.text.yview()[0])

    def load_template(self):

        self.set_code(
            "# Select a cell, apply its properties, then click Regenerate.\n"
        )

    def set_code(self, code):

        self.suppress_modified = True
        self.text.delete("1.0", "end")
        self.text.insert("1.0", code)
        self.text.edit_modified(False)
        self.suppress_modified = False
        self.update_line_numbers()

    def mark_clean(self):

        self.text.edit_modified(False)

    def get_code(self):

        return self.text.get("1.0", "end-1c")

    def import_style_csv(self):
        self._import_csv("style")

    def import_data_csv(self):
        self._import_csv("data")

    def _import_csv(self, target_var):
        import re
        import csv

        # Open file chooser
        file_path = filedialog.askopenfilename(
            title=f"Import {target_var.capitalize()} as CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            parent=self
        )
        if not file_path:
            return

        csv_rows = []
        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for row in reader:
                    csv_rows.append(row)
        except Exception:
            try:
                with open(file_path, "r", encoding="cp1251") as f:
                    reader = csv.reader(f)
                    for row in reader:
                        csv_rows.append(row)
            except Exception as exc:
                messagebox.showerror(
                    "Import Error",
                    f"Failed to read CSV file: {exc}",
                    parent=self
                )
                return

        if not csv_rows:
            messagebox.showwarning(
                "Empty File",
                "The selected CSV file contains no rows.",
                parent=self
            )
            return

        # Extract current code
        code = self.get_code()
        app = self.winfo_toplevel()

        try:
            preamble_str = app.extract_literal_assignment(code, "preamble")
            style_str = app.extract_literal_assignment(code, "style")
            data_str = app.extract_literal_assignment(code, "data")
        except Exception as exc:
            messagebox.showerror(
                "Parse Error",
                f"Cannot parse current code. Please fix any syntax errors first:\n{exc}",
                parent=self
            )
            return

        if preamble_str is None or style_str is None or data_str is None:
            messagebox.showerror(
                "Parse Error",
                "Cannot find literal assignments for preamble, style, or data in the code.",
                parent=self
            )
            return

        # Parse existing matrices
        def parse_matrix(matrix_str, var_name):
            inner = matrix_str.strip()
            prefix = f"{var_name}("
            if inner.lower().startswith(prefix.lower()):
                inner = inner[len(prefix):]
            if inner.endswith(")"):
                inner = inner[:-1]
            inner = inner.strip()

            raw_rows = inner.split(";")
            rows_list = []
            for r in raw_rows:
                r_clean = r.strip()
                if r_clean or len(raw_rows) > 1:
                    cells = [c.strip() for c in r_clean.split(",")]
                    rows_list.append(cells)
            return rows_list

        style_rows = parse_matrix(style_str, "style")
        data_rows = parse_matrix(data_str, "data")

        # Append to target matrix
        if target_var == "style":
            # Strip trailing empty rows from style so it aligns with active data
            while style_rows and all(c == "" for c in style_rows[-1]):
                style_rows.pop()
            style_rows.extend(csv_rows)
        else:
            # Strip trailing empty rows from data so it aligns with active style
            while data_rows and all(c == "" for c in data_rows[-1]):
                data_rows.pop()
            data_rows.extend(csv_rows)

        # Determine new dimensions
        new_rows = max(len(style_rows), len(data_rows))
        new_cols = max(
            max(len(r) for r in style_rows) if style_rows else 0,
            max(len(r) for r in data_rows) if data_rows else 0
        )

        # Re-serialize matrices
        def serialize_matrix(rows_list, var_name, target_rows, target_cols):
            # Pad columns
            for row in rows_list:
                while len(row) < target_cols:
                    row.append("")
            # Pad rows
            while len(rows_list) < target_rows:
                rows_list.append([""] * target_cols)

            serialized = []
            for row in rows_list:
                serialized.append(", ".join(row))
            return f"{var_name}(\n" + ";\n".join(serialized) + "\n)"

        new_style_val = serialize_matrix(style_rows, "style", new_rows, new_cols)
        new_data_val = serialize_matrix(data_rows, "data", new_rows, new_cols)

        # Update preamble dimensions
        m = re.match(r"^(\d+)\s*,\s*(\d+)(.*)$", preamble_str.strip())
        if not m:
            messagebox.showerror(
                "Preamble Error",
                "Failed to parse dimensions from preamble.",
                parent=self
            )
            return
        new_preamble_val = f"{new_rows},{new_cols}{m.group(3)}"

        # Replace back into code
        try:
            new_code = code
            new_code = app.replace_literal_assignment(new_code, "preamble", new_preamble_val)
            new_code = app.replace_literal_assignment(new_code, "style", new_style_val)
            new_code = app.replace_literal_assignment(new_code, "data", new_data_val)
        except Exception as exc:
            messagebox.showerror(
                "Replacement Error",
                f"Failed to update code: {exc}",
                parent=self
            )
            return

        # Update editor code
        self.set_code(new_code)
        # Trigger modification callback to update designer
        self._on_modified()
        messagebox.showinfo(
            "Success",
            f"Successfully imported and appended CSV rows to {target_var}.",
            parent=self
        )

class StatusBar(ttk.Frame):

    def __init__(self, master):

        super().__init__(master)

        self.pack(fill="x", side="bottom")

        self.label = ttk.Label(
            self,
            text="Ready",
            anchor="w"
        )

        self.label.pack(
            fill="x",
            padx=5,
            pady=2
        )

    def set(self, text):

        self.label.configure(text=text)


# ============================================================
# Main Window
# ============================================================

class TenTagsStudio(tk.Tk):

    def __init__(self):

        super().__init__()

        self.ui_font_family = choose_font_family(
            self,
            ("Segoe UI", "Arial", "DejaVu Sans", "TkDefaultFont")
        )
        self.editor_font_family = choose_font_family(
            self,
            (
                "Cascadia Mono",
                "Consolas",
                "DejaVu Sans Mono",
                "Courier New",
                "TkFixedFont"
            )
        )

        self.title("TenTags Studio")

        # Load and set application icon
        try:
            icon_path = APP_DIR / "tentags_studio_icon.png"
            if icon_path.is_file():
                from PIL import Image, ImageTk
                with Image.open(icon_path) as img:
                    photo = ImageTk.PhotoImage(img)
                    self.iconphoto(False, photo)
                    self._icon_ref = photo
        except Exception:
            pass

        self.geometry("1700x900")
        self.minsize(1300, 700)
        self.code_dirty = False
        self.designer_style_dirty = False
        self.designer_data_dirty = False
        self.synchronizing_model_to_designer = False
        self.regenerate_results = queue.Queue()

        self.toolbar = Toolbar(self)
        self.status = StatusBar(self)

        self.main = ttk.PanedWindow(
            self,
            orient=tk.HORIZONTAL
        )
        self.main.pack(fill="both", expand=True)

        self.workspace = tk.PanedWindow(
            self.main,
            orient=tk.VERTICAL,
            borderwidth=0,
            relief=tk.FLAT,
            sashwidth=7,
            sashrelief=tk.RAISED,
            sashcursor="sb_v_double_arrow",
            opaqueresize=True
        )

        self.designer = Designer(self.workspace)

        self.properties = PropertiesPanel(
            self.main,
            self.designer
        )

        self.cell_properties = CellProperties(
            self.main,
            self.designer
        )

        self.editor = CodeEditor(self.workspace)

        self.workspace.add(
            self.designer,
            minsize=250,
            stretch="always"
        )
        self.workspace.add(
            self.editor,
            minsize=120,
            stretch="always"
        )

        self.main.add(
            self.properties,
            weight=1
        )

        self.main.add(
            self.workspace,
            weight=6
        )

        self.main.add(
            self.cell_properties,
            weight=1
        )

        self.properties.set_status_callback(self.status.set)
        self.properties.set_before_resize_callback(self.prepare_grid_resize)
        self.properties.set_after_resize_callback(self.finish_grid_resize)
        self.properties.set_preview_callback(self.on_table_preview_changed)
        self.cell_properties.set_status_callback(self.status.set)
        self.cell_properties.set_style_change_callback(
            self.on_designer_style_changed
        )
        self.cell_properties.set_data_change_callback(
            self.on_designer_data_changed
        )
        self.cell_properties.set_output_change_callback(
            self.on_output_settings_changed
        )
        self.designer.set_edit_callback(self.on_cell_text_edited)

        self.toolbar.copy_btn.configure(
            command=self.copy_code
        )

        self.toolbar.regenerate_btn.configure(
            command=self.regenerate
        )

        self.toolbar.refresh_btn.configure(
            command=self.refresh_preview
        )

        self.editor.set_modified_callback(self.on_code_modified)
        self.after(100, self._poll_regenerate_results)
        self.after_idle(self._set_initial_workspace_split)

        self.load_default_document()

    def _set_initial_workspace_split(self):

        self.update_idletasks()
        height = self.workspace.winfo_height()
        if height > 0:
            self.workspace.sash_place(0, 0, int(height * 0.67))

    def load_default_document(self):

        model = tentags.compile(
            DEFAULT_PREAMBLE,
            DEFAULT_STYLE,
            DEFAULT_DATA
        )
        self.synchronize_model_to_designer(model, DEFAULT_PREAMBLE)
        self.editor.set_code(
            self.build_code(
                DEFAULT_PREAMBLE,
                DEFAULT_STYLE,
                DEFAULT_DATA
            )
        )
        self.designer_style_dirty = False
        self.designer_data_dirty = False
        self.mark_code_clean()
        self.status.set("Default 3 × 7 TenTags document loaded")

    def on_table_preview_changed(self, values):

        self.designer.set_preview_settings(
            cell_height=values["cell_height"],
            border_width=values["border_width"],
            border_color=values["border_color"],
            border_style=values["border_style"],
            row_scales=values["row_scales"],
            col_scales=values["col_scales"],
        )

        if self.synchronizing_model_to_designer:
            return
        if (
            values["rows"] != self.designer.rows
            or values["cols"] != self.designer.cols
        ):
            return

        try:
            code = self.editor.get_code()
            current_preamble = self.extract_literal_assignment(code, "preamble")
            if current_preamble is None or current_preamble == values["preamble"]:
                return
            code = self.replace_literal_assignment(
                code,
                "preamble",
                values["preamble"]
            )
            self.validate_edited_code(code)
        except Exception as exc:
            self.status.set(f"Grid settings are pending: {exc}")
            return

        self.editor.set_code(code)
        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set("Grid settings synchronized — click Regenerate")

    @staticmethod
    def _clean_attribute(value, replace_spaces=False):

        value = str(value).replace("\r", " ").replace("\n", " ")
        value = value.replace("<", "").replace(">", "").replace('"', "'")
        if replace_spaces:
            value = value.replace(" ", "%20")
        return value

    @staticmethod
    def _data_text(value):

        value = str(value).replace("\r\n", "\n").replace("\r", "\n")
        value = value.replace("\n", " ").replace('"', "'")
        if "," in value or ";" in value:
            return f'"{value}"'
        return value

    def _build_compact_style(self):

        states = [
            self.designer.states[(row, col)]
            for row in range(self.designer.rows)
            for col in range(self.designer.cols)
        ]
        if not states:
            return "style()"

        layers = [
            (
                lambda state: state.align.lower(),
                lambda value: f"<{value}>",
                lambda value: f"</{value}>",
            ),
            (
                lambda state: state.bg,
                lambda value: f"<bg={value}>",
                lambda value: "</bg>",
            ),
            (
                lambda state: state.fg,
                lambda value: f"<color={value}>",
                lambda value: "</color>",
            ),
            (
                lambda state: state.bold,
                lambda value: "<b>" if value else "",
                lambda value: "</b>" if value else "",
            ),
            (
                lambda state: state.italic,
                lambda value: "<i>" if value else "",
                lambda value: "</i>" if value else "",
            ),
            (
                lambda state: state.underline,
                lambda value: "<u>" if value else "",
                lambda value: "</u>" if value else "",
            ),
            (
                lambda state: state.strike,
                lambda value: "<s>" if value else "",
                lambda value: "</s>" if value else "",
            ),
            (
                lambda state: state.font_size,
                lambda value: (
                    f"<fs={self._clean_attribute(value)}>" if value else ""
                ),
                lambda value: "</fs>" if value else "",
            ),
        ]

        prefixes = [[] for _ in states]
        suffixes = [[] for _ in states]

        def encode_layer(layer_index, start, end):
            if layer_index >= len(layers):
                return

            value_getter, opening_tag, closing_tag = layers[layer_index]
            cursor = start

            while cursor <= end:
                value = value_getter(states[cursor])
                group_end = cursor
                while (
                    group_end < end
                    and value_getter(states[group_end + 1]) == value
                ):
                    group_end += 1

                opening = opening_tag(value)
                closing = closing_tag(value)
                if opening:
                    prefixes[cursor].append(opening)

                encode_layer(layer_index + 1, cursor, group_end)

                if closing:
                    suffixes[group_end].append(closing)
                cursor = group_end + 1

        encode_layer(0, 0, len(states) - 1)
        expressions = [
            "".join(prefixes[index] + suffixes[index])
            for index in range(len(states))
        ]
        rows = [
            ", ".join(expressions[row * self.designer.cols:(row + 1) * self.designer.cols])
            for row in range(self.designer.rows)
        ]
        return "style(\n" + ";\n".join(rows) + "\n)"

    def _data_expression(self, state):

        parts = []

        if state.mark:
            mark = self._clean_attribute(state.mark).replace(" ", "_")
            parts.append(f"<mark={mark}>")

        if state.image:
            image_path = Path(state.image)
            if image_path.is_absolute():
                try:
                    image_path = image_path.resolve().relative_to(APP_DIR)
                except ValueError:
                    pass

            source = self._clean_attribute(
                str(image_path).replace("\\", "/"),
                replace_spaces=True
            )
            width = self._clean_attribute(state.image_w) or "120"
            height = self._clean_attribute(state.image_h) or "auto"
            margin = self._clean_attribute(state.image_m) or "0"
            parts.append(f"<img src={source} w={width} h={height} m={margin}>")

        if state.text:
            text = self._data_text(state.text)
            if state.url:
                url = self._clean_attribute(state.url, replace_spaces=True)
                text = f"<url={url}>{text}</url>"
            parts.append(text)

        return "".join(parts) or " "

    def build_blocks(self):

        data_grid = []

        for row in range(self.designer.rows):
            data_row = []

            for col in range(self.designer.cols):
                state = self.designer.states[(row, col)]
                data_row.append(self._data_expression(state))

            data_grid.append(data_row)

        for (row, col), state in self.designer.states.items():
            if state.hide_top or state.hide_bottom:
                data_grid[row][col] = f"<rm>{data_grid[row][col]}</rm>"

        for row in range(self.designer.rows):
            col = 0
            while col < self.designer.cols - 1:
                if not self.designer.states[(row, col)].hide_right:
                    col += 1
                    continue

                start_col = col
                end_col = col + 1
                while (
                    end_col < self.designer.cols - 1
                    and self.designer.states[(row, end_col)].hide_right
                ):
                    end_col += 1

                data_grid[row][start_col] = "<cm>" + data_grid[row][start_col]
                data_grid[row][end_col] += "</cm>"
                col = end_col + 1

        data_rows = [", ".join(row) for row in data_grid]

        style = self._build_compact_style()
        data = "data(\n" + ";\n".join(data_rows) + "\n)"
        return style, data

    @staticmethod
    def _python_triple_string(value):

        escaped = value.replace("\\", "\\\\").replace('"""', '\\"\\"\\"')
        return f'"""{escaped}"""'

    def build_output_code(self):

        settings = self.cell_properties.get_output_settings()
        lines = [
            "# <tentags-output>",
            "import os",
            "from pathlib import Path",
            "",
            '_proj_root = os.environ.get("TENTAGS_PROJECT_ROOT")',
            "if _proj_root:",
            '    _export_dir = Path(_proj_root) / "export_files"',
            "else:",
            "    _file_dir = Path(__file__).resolve().parent",
            '    if (_file_dir / "export_files").is_dir():',
            '        _export_dir = _file_dir / "export_files"',
            '    elif (_file_dir.parent / "export_files").is_dir():',
            '        _export_dir = _file_dir.parent / "export_files"',
            "    else:",
            "        _export_dir = _file_dir",
            "os.makedirs(_export_dir, exist_ok=True)",
        ]

        if settings["html"]:
            lines.append("")
            if settings["html_full_page"]:
                lines.extend([
                    "import html as _html",
                    "",
                    "html_table = tentags.render_html(model)",
                    f'html_title = {settings["html_title"]!r}',
                    "html_document = (",
                    '    "<!DOCTYPE html><html><head><meta charset=\\"utf-8\\">"',
                    '    f"<title>{_html.escape(html_title)}</title></head><body>"',
                    "    + html_table",
                    '    + "</body></html>"',
                    ")",
                    'html_path = _export_dir / "report.html"',
                    'with open(html_path, "w", encoding="utf-8") as f:',
                    "    f.write(html_document)",
                ])
            else:
                lines.extend([
                    'html_path = _export_dir / "report.html"',
                    'with open(html_path, "w", encoding="utf-8") as f:',
                    "    f.write(tentags.render_html(model))",
                ])

        if settings["pdf"]:
            pdf_settings = {
                "page_size": settings["pdf_page_size"],
                "orientation": settings["pdf_orientation"],
                "margins": settings["pdf_margins"],
            }
            lines.extend([
                "",
                "from studio_renderers import render_pdf_with_images as _render_pdf",
                "",
                f"pdf_settings = {pdf_settings!r}",
                'pdf_path = _export_dir / "report.pdf"',
                '_render_pdf(model, str(pdf_path), settings=pdf_settings)',
            ])

        if settings["xlsx"]:
            paper_codes = {
                "letter": "1",
                "legal": "5",
                "tabloid": "3",
                "A3": "8",
                "A4": "9",
                "A5": "11",
            }
            lines.extend([
                "",
                "from openpyxl import load_workbook as _load_workbook",
                "from studio_renderers import render_xlsx_with_images as _render_xlsx",
                "",
                'xlsx_path = _export_dir / "report.xlsx"',
                '_render_xlsx(model, str(xlsx_path))',
                'workbook = _load_workbook(str(xlsx_path))',
                "worksheet = workbook.active",
                f'worksheet.title = {settings["xlsx_sheet_name"]!r}',
                f'worksheet.page_setup.orientation = {settings["xlsx_orientation"]!r}',
                f"worksheet.page_setup.paperSize = "
                f"{paper_codes[settings['xlsx_page_size']]!r}",
                f"worksheet.sheet_view.showGridLines = "
                f"{settings['xlsx_gridlines']!r}",
                f"worksheet.sheet_properties.pageSetUpPr.fitToPage = "
                f"{settings['xlsx_fit_to_page']!r}",
            ])
            if settings["xlsx_fit_to_page"]:
                lines.extend([
                    "worksheet.page_setup.fitToWidth = 1",
                    "worksheet.page_setup.fitToHeight = 0",
                ])
            lines.append('workbook.save(str(xlsx_path))')

        lines.append("# </tentags-output>")
        return "\n".join(lines)

    def build_code(self, preamble, style, data):

        lines = [
            "# -*- coding: utf-8 -*-",
            "",
            "import tentags",
            "",
            f"preamble = {preamble!r}",
            "",
            "# Compact starter template. Move its tags to any cells you want.",
            "# Regenerate preserves their placement and applies it to the designer.",
            f"style = {self._python_triple_string(style)}",
            "",
            f"data = {self._python_triple_string(data)}",
            "",
            "model = tentags.compile(preamble, style, data)",
            "",
            self.build_output_code(),
        ]

        return "\n".join(lines) + "\n"

    @staticmethod
    def extract_literal_assignment(code, name):

        tree = ast.parse(code, filename="<generated_tentags>", mode="exec")
        assignment_found = False

        for node in tree.body:
            targets = []
            value_node = None

            if isinstance(node, ast.Assign):
                targets = node.targets
                value_node = node.value
            elif isinstance(node, ast.AnnAssign):
                targets = [node.target]
                value_node = node.value

            if not any(
                isinstance(target, ast.Name) and target.id == name
                for target in targets
            ):
                continue

            assignment_found = True
            try:
                value = ast.literal_eval(value_node)
            except (ValueError, TypeError):
                continue

            if isinstance(value, str):
                return value

        if assignment_found:
            raise ValueError(
                f"TenTags Studio cannot safely preserve dynamic {name}. "
                f"Use a literal string assignment for {name}."
            )
        return None

    @staticmethod
    def style_matches_dimensions(style, rows, cols):

        style_model = tentags.compile("", style, "data()")
        return (
            len(style_model.cells) == rows
            and all(len(row) == cols for row in style_model.cells)
        )

    @classmethod
    def replace_literal_assignment(cls, code, name, value):

        tree = ast.parse(code, filename="<edited_tentags>", mode="exec")
        value_nodes = []

        for node in tree.body:
            targets = []
            value_node = None
            if isinstance(node, ast.Assign):
                targets = node.targets
                value_node = node.value
            elif isinstance(node, ast.AnnAssign):
                targets = [node.target]
                value_node = node.value

            if any(
                isinstance(target, ast.Name) and target.id == name
                for target in targets
            ):
                value_nodes.append(value_node)

        if not value_nodes:
            raise ValueError(f"Cannot find the {name} assignment to resize it.")

        value_node = value_nodes[-1]
        lines = code.splitlines(keepends=True)

        def absolute_offset(line_number, byte_column):
            line = lines[line_number - 1]
            prefix = line.encode("utf-8")[:byte_column].decode(
                "utf-8",
                errors="ignore"
            )
            return sum(len(item) for item in lines[:line_number - 1]) + len(prefix)

        start = absolute_offset(value_node.lineno, value_node.col_offset)
        end = absolute_offset(value_node.end_lineno, value_node.end_col_offset)
        literal = cls._python_triple_string(value)
        return code[:start] + literal + code[end:]

    def prepare_grid_resize(self, values):

        if (
            self.designer.current_cell
            and not self.cell_properties.apply_before_selection()
        ):
            return False

        if self.code_dirty:
            code = self.editor.get_code()
            try:
                self.validate_edited_code(code)
                model = self.compile_editor_model(code)
                if model is not None:
                    self.synchronize_model_to_designer(
                        model,
                        self.extract_literal_assignment(code, "preamble")
                    )
                    self.designer_style_dirty = False
            except SyntaxError as exc:
                self.status.set("Cannot resize — fix the lower editor syntax first")
                messagebox.showerror(
                    "Syntax error",
                    self.format_syntax_error(exc),
                    parent=self
                )
                return False
            except Exception as exc:
                self.status.set("Cannot resize — invalid TenTags code")
                messagebox.showerror("TenTags syntax error", str(exc), parent=self)
                return False

        if self.designer.current_cell:
            self._grid_resize_selection = (
                self.designer.current_cell.row,
                self.designer.current_cell.col,
            )
        else:
            self._grid_resize_selection = (0, 0)
        return True

    def finish_grid_resize(self, values):

        rows = values["rows"]
        cols = values["cols"]
        self.properties.set_dimensions(rows, cols)

        selected_row, selected_col = self._grid_resize_selection
        self.designer.select(
            min(selected_row, rows - 1),
            min(selected_col, cols - 1)
        )

        preamble = self.properties.get_preamble()
        style, data = self.build_blocks()
        code = self.editor.get_code()
        code = self.replace_literal_assignment(code, "preamble", preamble)
        code = self.replace_literal_assignment(code, "style", style)
        code = self.replace_literal_assignment(code, "data", data)

        tentags.compile(preamble, style, data)
        self.validate_edited_code(code)
        self.editor.set_code(code)
        self.designer_style_dirty = False
        self.designer_data_dirty = False
        self.code_dirty = True
        self.toolbar.show_regenerate()

    def code_with_designer_data(self, code):

        _, data = self.build_blocks()
        updated_code = self.replace_literal_assignment(code, "data", data)
        self.validate_edited_code(updated_code)
        self.compile_editor_model(updated_code)
        return updated_code

    def code_with_designer_style(self, code):

        style = self._build_compact_style()
        updated_code = self.replace_literal_assignment(code, "style", style)
        self.validate_edited_code(updated_code)
        self.compile_editor_model(updated_code)
        return updated_code

    def code_with_output_settings(self, code):

        start_marker = "# <tentags-output>"
        end_marker = "# </tentags-output>"
        output_code = self.build_output_code()
        start = code.find(start_marker)
        end = code.find(end_marker, start + len(start_marker))

        if start == -1 or end == -1:
            updated_code = code.rstrip() + "\n\n" + output_code + "\n"
        else:
            end += len(end_marker)
            updated_code = code[:start] + output_code + code[end:]

        self.validate_edited_code(updated_code)
        self.compile_editor_model(updated_code)
        return updated_code

    @staticmethod
    def split_source_grid(block, block_name):

        prefix = f"{block_name}("
        stripped = block.strip()
        if not stripped.lower().startswith(prefix.lower()) or not stripped.endswith(")"):
            raise ValueError(f"Expected a {block_name}(...) block.")

        content = stripped[len(prefix):-1]
        rows = [[]]
        buffer = []
        quote = None
        escaped = False
        inside_tag = False

        def commit_cell():
            rows[-1].append("".join(buffer).strip())
            buffer.clear()

        for character in content:
            if quote is not None:
                buffer.append(character)
                if escaped:
                    escaped = False
                elif character == "\\":
                    escaped = True
                elif character == quote:
                    quote = None
                continue

            if inside_tag:
                buffer.append(character)
                if character == ">":
                    inside_tag = False
                continue

            if character in {'"', "'"}:
                quote = character
                buffer.append(character)
            elif character == "<":
                inside_tag = True
                buffer.append(character)
            elif character == ",":
                commit_cell()
            elif character == ";":
                commit_cell()
                rows.append([])
            else:
                buffer.append(character)

        commit_cell()
        return rows

    @staticmethod
    def replace_expression_text(expression, replacement):

        segments = []
        buffer = []
        inside_tag = False

        def commit_segment(is_tag):
            if buffer:
                segments.append((is_tag, "".join(buffer)))
                buffer.clear()

        for character in expression:
            if not inside_tag and character == "<":
                commit_segment(False)
                inside_tag = True
                buffer.append(character)
            elif inside_tag and character == ">":
                buffer.append(character)
                commit_segment(True)
                inside_tag = False
            else:
                buffer.append(character)
        commit_segment(inside_tag)

        replaced = False
        result = []
        for is_tag, value in segments:
            if not is_tag and value.strip() and not replaced:
                leading = value[:len(value) - len(value.lstrip())]
                trailing = value[len(value.rstrip()):]
                result.append(leading + replacement + trailing)
                replaced = True
            elif not is_tag and value.strip() and replaced:
                continue
            else:
                result.append(value)

        if not replaced:
            insert_at = next(
                (
                    index
                    for index, value in enumerate(result)
                    if value.lstrip().startswith("</")
                ),
                len(result)
            )
            result.insert(insert_at, replacement)
        return "".join(result)

    def code_with_designer_cell_text(self, code, cell):

        data = self.extract_literal_assignment(code, "data")
        rows = self.split_source_grid(data, "data")
        if cell.row >= len(rows) or cell.col >= len(rows[cell.row]):
            raise ValueError("The selected cell is outside the data(...) grid.")

        replacement = self._data_text(cell.state.text) if cell.state.text else " "
        rows[cell.row][cell.col] = self.replace_expression_text(
            rows[cell.row][cell.col],
            replacement
        )
        updated_data = "data(\n" + ";\n".join(
            ", ".join(row) for row in rows
        ) + "\n)"
        updated_code = self.replace_literal_assignment(code, "data", updated_data)
        self.validate_edited_code(updated_code)
        self.compile_editor_model(updated_code)
        return updated_code

    def generate(self, execute=True):

        if self.code_dirty:
            self.regenerate()
            return

        if (
            self.designer.current_cell
            and not self.cell_properties.apply_before_selection()
        ):
            return

        try:
            values = self.properties.get_values()
            target_rows = values["rows"]
            target_cols = values["cols"]

            if (
                self.designer.rows != target_rows
                or self.designer.cols != target_cols
            ):
                self.designer.resize(target_rows, target_cols)

            preamble = self.properties.get_preamble()
            generated_style, generated_data = self.build_blocks()
            current_data = self.extract_literal_assignment(
                self.editor.get_code(),
                "data"
            )
            data = current_data if current_data is not None else generated_data
            current_style = self.extract_literal_assignment(
                self.editor.get_code(),
                "style"
            )
            style_was_resized = bool(
                current_style is not None
                and not self.style_matches_dimensions(
                    current_style,
                    target_rows,
                    target_cols
                )
            )
            style = (
                generated_style
                if (
                    current_style is None
                    or style_was_resized
                    or self.designer_style_dirty
                )
                else current_style
            )

            tentags.compile(preamble, style, data)
            code = self.build_code(preamble, style, data)
            compile(code, "<generated_tentags>", "exec")

            self.editor.set_code(code)
            style_was_edited_in_panel = self.designer_style_dirty
            self.designer_style_dirty = False
            self.designer_data_dirty = False
            self.mark_code_clean()
            status = (
                f"Code generated for {self.designer.rows} × "
                f"{self.designer.cols} table"
            )
            if style_was_resized:
                status += " — style grid synchronized with preamble"
            elif style_was_edited_in_panel:
                status += " — cell styles applied"
            elif current_style is not None:
                status += " — manual style preserved"
            self.status.set(status)

            if execute:
                self.regenerate()
        except Exception as exc:
            messagebox.showerror("Generation error", str(exc), parent=self)
            self.status.set("Generation failed")

    def copy_code(self):

        code = self.editor.get_code()
        self.clipboard_clear()
        self.clipboard_append(code)
        self.update_idletasks()
        self.status.set("Generated code copied to clipboard")

    def on_code_modified(self):

        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set(
            "Manual code changes detected — click Regenerate to run them"
        )

    def on_designer_style_changed(self):

        try:
            code = self.code_with_designer_style(self.editor.get_code())
        except Exception as exc:
            self.designer_style_dirty = True
            self.status.set(f"Cell style is pending: {exc}")
            return

        self.editor.set_code(code)
        self.designer_style_dirty = False
        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set(
            "Cell style synchronized — click Regenerate to create the reports"
        )

    def on_output_settings_changed(self):

        try:
            code = self.code_with_output_settings(self.editor.get_code())
        except Exception as exc:
            self.status.set(f"Output settings are pending: {exc}")
            raise

        self.editor.set_code(code)
        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set("Output settings synchronized — click Regenerate")

    def on_designer_data_changed(self):

        self.designer_data_dirty = True
        try:
            code = self.code_with_designer_data(self.editor.get_code())
        except Exception:
            self.status.set(
                "Cell data changed — fix the lower editor syntax before Regenerate"
            )
            return

        self.editor.set_code(code)
        self.designer_data_dirty = False
        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set("Cell data synchronized — click Regenerate")

    def on_cell_text_edited(self, cell):

        try:
            code = self.code_with_designer_cell_text(
                self.editor.get_code(),
                cell
            )
        except Exception as exc:
            self.designer_data_dirty = True
            self.status.set(f"Cell text is pending: {exc}")
            return

        self.editor.set_code(code)
        self.designer_data_dirty = False
        self.code_dirty = True
        self.toolbar.show_regenerate()
        self.status.set(
            f"Cell {self.designer.cell_name(cell.row, cell.col)} edited — "
            "lower data synchronized"
        )

    def mark_code_clean(self):

        self.code_dirty = False
        self.editor.mark_clean()
        self.toolbar.hide_regenerate()

    @staticmethod
    def validate_edited_code(code):

        return compile(code, "<edited_tentags>", "exec")

    @staticmethod
    def extract_tentags_blocks(code):

        tree = ast.parse(code, filename="<edited_tentags>", mode="exec")
        names = {"preamble", "style", "data"}
        assigned = set()
        values = {}

        for node in tree.body:
            targets = []
            value_node = None

            if isinstance(node, ast.Assign):
                targets = node.targets
                value_node = node.value
            elif isinstance(node, ast.AnnAssign):
                targets = [node.target]
                value_node = node.value

            target_names = [
                target.id
                for target in targets
                if isinstance(target, ast.Name) and target.id in names
            ]

            for name in target_names:
                assigned.add(name)
                try:
                    value = ast.literal_eval(value_node)
                except (ValueError, TypeError):
                    values.pop(name, None)
                    continue

                if isinstance(value, str):
                    values[name] = value

        if not assigned:
            return None

        missing = names - values.keys()
        if missing:
            missing_list = ", ".join(sorted(missing))
            raise ValueError(
                "Visual synchronization requires literal string assignments for "
                f"preamble, style and data. Missing or dynamic: {missing_list}."
            )

        return values["preamble"], values["style"], values["data"]

    def compile_editor_model(self, code):

        blocks = self.extract_tentags_blocks(code)
        if blocks is None:
            return None
        return tentags.compile(*blocks)

    @staticmethod
    def _model_merge_groups(model):

        horizontal = []
        vertical = []

        def has_flag(row, col, flag):
            return (
                row < len(model.cells)
                and col < len(model.cells[row])
                and bool(model.cells[row][col].border_flags & flag)
            )

        for row in range(model.rows):
            col = 0
            while col < model.cols - 1:
                if not has_flag(row, col, tentags.BorderFlags.HIDE_RIGHT):
                    col += 1
                    continue

                start = col
                end = col
                while (
                    end < model.cols - 1
                    and has_flag(row, end, tentags.BorderFlags.HIDE_RIGHT)
                ):
                    end += 1

                horizontal.append((row, start, end - start + 1))
                col = end + 1

        for col in range(model.cols):
            row = 0
            while row < model.rows - 1:
                if not has_flag(row, col, tentags.BorderFlags.HIDE_BOTTOM):
                    row += 1
                    continue

                start = row
                end = row
                while (
                    end < model.rows - 1
                    and has_flag(end, col, tentags.BorderFlags.HIDE_BOTTOM)
                ):
                    end += 1

                vertical.append((start, col, end - start + 1))
                row = end + 1

        return horizontal, vertical

    def _preview_color(self, value, default):

        color = str(value or default)
        try:
            self.winfo_rgb(color)
        except tk.TclError:
            return default
        return color

    def _apply_model_cell(self, state, cell):

        styles = cell.styles
        align = str(styles.get("text-align", "center")).lower()

        state.text = str(cell.raw_expr or "")
        state.bold = str(styles.get("font-weight", "")).lower() == "bold"
        state.italic = str(styles.get("font-style", "")).lower() == "italic"
        decorations = str(styles.get("text-decoration", "")).lower().split()
        state.underline = "underline" in decorations
        state.strike = "line-through" in decorations
        state.font_size = str(styles.get("font-size", ""))
        state.align = {
            "left": "Left",
            "center": "Center",
            "right": "Right",
        }.get(align, "Center")
        state.fg = self._preview_color(styles.get("color"), "#000000")
        state.bg = self._preview_color(
            styles.get("background-color"),
            "#ffffff"
        )

        state.url = str(cell.link.raw) if cell.link is not None else ""
        state.mark = str(cell.mark or "")

        if cell.images:
            image = cell.images[0]
            state.image = str(image.get("src", ""))
            state.image_w = str(image.get("w", "120"))
            state.image_h = str(image.get("h", "auto"))
            state.image_m = str(image.get("m", "0"))
        else:
            state.image = ""
            state.image_w = "120"
            state.image_h = "auto"
            state.image_m = "15"

        flags = cell.border_flags
        state.hide_left = bool(flags & tentags.BorderFlags.HIDE_LEFT)
        state.hide_right = bool(flags & tentags.BorderFlags.HIDE_RIGHT)
        state.hide_top = bool(flags & tentags.BorderFlags.HIDE_TOP)
        state.hide_bottom = bool(flags & tentags.BorderFlags.HIDE_BOTTOM)

    @staticmethod
    def _scale_text_from_preamble(preamble):

        marker = "scale("
        start = preamble.lower().find(marker)
        if start == -1:
            return ""
        start += len(marker)
        end = preamble.rfind(")")
        if end < start:
            return ""
        return preamble[start:end].strip()

    def synchronize_model_to_designer(self, model, preamble=None):

        horizontal, vertical = self._model_merge_groups(model)

        selected_row = 0
        selected_col = 0
        if self.designer.current_cell:
            selected_row = self.designer.current_cell.row
            selected_col = self.designer.current_cell.col

        self.synchronizing_model_to_designer = True
        try:
            scale_text = (
                self._scale_text_from_preamble(preamble)
                if preamble is not None
                else None
            )
            self.properties.load_model_settings(model, scale_text=scale_text)
        finally:
            self.synchronizing_model_to_designer = False
        self.designer.resize(model.rows, model.cols)

        for row in range(model.rows):
            for col in range(model.cols):
                state = self.designer.states[(row, col)]
                if row < len(model.cells) and col < len(model.cells[row]):
                    self._apply_model_cell(state, model.cells[row][col])
                else:
                    self._apply_model_cell(state, tentags.CellDesc())

        self.designer.refresh_grid_borders()

        selected_row = min(selected_row, model.rows - 1)
        selected_col = min(selected_col, model.cols - 1)
        self.designer.select(selected_row, selected_col)

        return len(horizontal), len(vertical)

    @staticmethod
    def format_syntax_error(error):

        location = f"Line {error.lineno or '?'}, column {error.offset or '?'}"
        source_line = (error.text or "").rstrip()
        if source_line:
            return f"{error.msg}\n\n{location}\n{source_line}"
        return f"{error.msg}\n\n{location}"

    def _apply_editor_code_to_preview(self, code):

        if self.designer_data_dirty:
            code = self.code_with_designer_data(code)
            self.designer_data_dirty = False

        model = self.compile_editor_model(code)
        if model is None:
            raise ValueError(
                "Cannot refresh preview without literal preamble, style and data."
            )

        horizontal_count, vertical_count = self.synchronize_model_to_designer(
            model,
            self.extract_literal_assignment(code, "preamble")
        )
        self.designer_style_dirty = False
        self.designer_data_dirty = False

        current_style = self.extract_literal_assignment(code, "style")
        if not self.style_matches_dimensions(
            current_style,
            model.rows,
            model.cols
        ):
            resized_style = self._build_compact_style()
            code = self.replace_literal_assignment(
                code,
                "style",
                resized_style
            )
            self.compile_editor_model(code)

        if self.editor.get_code() != code:
            self.editor.set_code(code)
        return code, horizontal_count, vertical_count

    def refresh_preview(self):

        code = self.editor.get_code()
        try:
            self.validate_edited_code(code)
        except SyntaxError as exc:
            self.status.set("Python syntax error — preview was not refreshed")
            messagebox.showerror(
                "Syntax error",
                self.format_syntax_error(exc),
                parent=self
            )
            return

        try:
            _, horizontal_count, vertical_count = (
                self._apply_editor_code_to_preview(code)
            )
        except Exception as exc:
            self.status.set("TenTags syntax error — preview was not refreshed")
            messagebox.showerror(
                "TenTags syntax error",
                str(exc),
                parent=self
            )
            return

        self.status.set(
            "Preview refreshed without generating files — synchronized "
            f"{horizontal_count} <cm> and {vertical_count} <rm> merges"
        )

    def regenerate(self):

        code = self.editor.get_code()

        try:
            self.validate_edited_code(code)
        except SyntaxError as exc:
            self.status.set("Python syntax error — changes were not executed")
            messagebox.showerror(
                "Syntax error",
                self.format_syntax_error(exc),
                parent=self
            )
            return

        try:
            code, horizontal_count, vertical_count = (
                self._apply_editor_code_to_preview(code)
            )
            self.status.set(
                "Syntax OK — editor applied to designer; synchronized "
                f"{horizontal_count} <cm> and {vertical_count} <rm> merges"
            )
        except Exception as exc:
            self.status.set("TenTags syntax error — changes were not executed")
            messagebox.showerror(
                "TenTags syntax error",
                str(exc),
                parent=self
            )
            return

        self.toolbar.regenerate_btn.configure(
            state="disabled",
            text="Regenerating..."
        )
        self.status.set("Syntax OK — executing edited code...")

        worker = threading.Thread(
            target=self._execute_edited_code,
            args=(code,),
            daemon=True
        )
        worker.start()

    def _execute_edited_code(self, code):

        script_path = None

        try:
            with tempfile.NamedTemporaryFile(
                mode="w",
                encoding="utf-8",
                suffix=".py",
                delete=False
            ) as script:
                script.write(code)
                script_path = Path(script.name)

            creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
            process_env = os.environ.copy()
            process_env["PYTHONUTF8"] = "1"
            process_env["PYTHONIOENCODING"] = "utf-8"
            process_env["TENTAGS_PROJECT_ROOT"] = str(APP_DIR.parent)
            existing_pythonpath = process_env.get("PYTHONPATH")
            parent_dir = str(APP_DIR.parent)
            process_env["PYTHONPATH"] = (
                f"{APP_DIR}{os.pathsep}{parent_dir}"
                if not existing_pythonpath
                else f"{APP_DIR}{os.pathsep}{parent_dir}{os.pathsep}{existing_pythonpath}"
            )

            result = subprocess.run(
                [sys.executable, str(script_path)],
                cwd=str(APP_DIR),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=120,
                creationflags=creation_flags,
                env=process_env
            )

            if result.returncode == 0:
                details = result.stdout.strip()
                self.regenerate_results.put((True, details, code))
            else:
                details = result.stderr.strip() or result.stdout.strip()
                self.regenerate_results.put((False, details, code))

        except subprocess.TimeoutExpired:
            self.regenerate_results.put((
                False,
                "Execution exceeded the 120 second timeout.",
                code
            ))
        except Exception as exc:
            self.regenerate_results.put((False, str(exc), code))
        finally:
            if script_path:
                script_path.unlink(missing_ok=True)

    def _poll_regenerate_results(self):

        try:
            while True:
                result = self.regenerate_results.get_nowait()
                self._finish_regenerate(*result)
        except queue.Empty:
            pass

        if self.winfo_exists():
            self.after(100, self._poll_regenerate_results)

    def _finish_regenerate(self, success, details, executed_code):

        self.toolbar.regenerate_btn.configure(
            state="normal",
            text="Regenerate"
        )

        if success:
            if self.editor.get_code() == executed_code:
                self.mark_code_clean()
                self.status.set("Edited code executed successfully")
            else:
                self.code_dirty = True
                self.toolbar.show_regenerate()
                self.status.set("Previous version executed; newer changes are pending")

            message = "Syntax is valid and the edited code executed successfully."
            if details:
                message += f"\n\nOutput:\n{details[:2000]}"
            messagebox.showinfo("Regenerate complete", message, parent=self)
        else:
            self.code_dirty = True
            self.toolbar.show_regenerate()
            self.status.set("Edited code failed during execution")
            messagebox.showerror(
                "Regenerate failed",
                (details or "Unknown execution error")[:5000],
                parent=self
            )


# ============================================================
# main
# ============================================================

def main():

    app = TenTagsStudio()
    app.mainloop()


if __name__ == "__main__":
    main()
